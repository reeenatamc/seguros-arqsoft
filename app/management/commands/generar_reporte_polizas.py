import os
from datetime import datetime, timedelta
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import Factura, Pago, Poliza


class Command(BaseCommand):
    help = "Genera reportes de pólizas en formato Excel y PDF"

    def add_arguments(self, parser):
        parser.add_argument(
            "--formato",
            type=str,
            choices=["excel", "pdf", "ambos"],
            default="ambos",
            help="Formato del reporte: excel, pdf o ambos",
        )
        parser.add_argument(
            "--estado",
            type=str,
            choices=["vigente", "vencida", "cancelada", "por_vencer", "todas"],
            default="todas",
            help="Filtrar pólizas por estado",
        )

    def handle(self, *args, **options):
        formato = options["formato"]
        estado = options["estado"]

        self.stdout.write(self.style.SUCCESS(f"Generando reporte de pólizas..."))

        # OPTIMIZACIÓN: usar select_related y prefetch_related para evitar N+1
        base_queryset = Poliza.objects.select_related(
            "compania_aseguradora", "corredor_seguros", "tipo_poliza"
        ).prefetch_related("facturas", Prefetch("facturas__pagos", queryset=Pago.objects.filter(estado="aprobado")))

        # Obtener pólizas según filtro
        if estado == "todas":
            polizas = base_queryset.all()
        else:
            polizas = base_queryset.filter(estado=estado)

        # Crear directorio de reportes si no existe
        reportes_dir = "media/reportes/polizas"
        os.makedirs(reportes_dir, exist_ok=True)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        if formato in ["excel", "ambos"]:
            excel_file = f"{reportes_dir}/reporte_polizas_{estado}_{timestamp}.xlsx"
            self.generar_excel(polizas, excel_file, estado)
            self.stdout.write(self.style.SUCCESS(f"  ✓ Reporte Excel generado: {excel_file}"))

        if formato in ["pdf", "ambos"]:
            pdf_file = f"{reportes_dir}/reporte_polizas_{estado}_{timestamp}.pdf"
            self.generar_pdf(polizas, pdf_file, estado)
            self.stdout.write(self.style.SUCCESS(f"  ✓ Reporte PDF generado: {pdf_file}"))

        self.stdout.write(self.style.SUCCESS(f"✓ Reporte completado: {polizas.count()} pólizas procesadas"))

    def generar_excel(self, polizas, filename, estado):
        """Genera el reporte en formato Excel"""
        wb = openpyxl.Workbook()

        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self.crear_hoja_resumen(ws_resumen, polizas)

        # Hoja 2: Detalle de Pólizas
        ws_detalle = wb.create_sheet("Detalle de Pólizas")
        self.crear_hoja_detalle_polizas(ws_detalle, polizas)

        # Hoja 3: Gastos por Póliza
        ws_gastos = wb.create_sheet("Gastos por Póliza")
        self.crear_hoja_gastos(ws_gastos, polizas)

        wb.save(filename)

    def crear_hoja_resumen(self, ws, polizas):
        """Crea la hoja de resumen del reporte - OPTIMIZADO con agregaciones SQL"""
        # Título
        ws["A1"] = "REPORTE DE PÓLIZAS - RESUMEN EJECUTIVO"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")

        # Fecha de generación
        ws["A2"] = f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws["A2"].font = Font(size=10, italic=True)

        # Resumen por estado
        ws["A4"] = "RESUMEN POR ESTADO"
        ws["A4"].font = Font(size=12, bold=True)

        headers = ["Estado", "Cantidad", "Suma Asegurada Total", "Porcentaje"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # OPTIMIZACIÓN: una sola consulta para todos los estados
        resumen_estados = polizas.values("estado").annotate(
            cantidad=Count("id"), suma_total=Coalesce(Sum("suma_asegurada"), Decimal("0"))
        )

        # Convertir a diccionario para acceso rápido
        estados_dict = {r["estado"]: r for r in resumen_estados}
        total_polizas = sum(r["cantidad"] for r in resumen_estados)

        estados_orden = ["vigente", "vencida", "por_vencer", "cancelada"]
        row = 6

        for estado in estados_orden:
            data = estados_dict.get(estado, {"cantidad": 0, "suma_total": Decimal("0")})
            cantidad = data["cantidad"]
            suma_total = data["suma_total"]
            porcentaje = (cantidad / total_polizas * 100) if total_polizas > 0 else 0

            ws.cell(row=row, column=1, value=estado.replace("_", " ").title())
            ws.cell(row=row, column=2, value=cantidad)
            ws.cell(row=row, column=3, value=f"${suma_total:,.2f}")
            ws.cell(row=row, column=4, value=f"{porcentaje:.1f}%")
            row += 1

        # Total
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_polizas)
        ws.cell(row=row, column=2).font = Font(bold=True)

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15

    def crear_hoja_detalle_polizas(self, ws, polizas):
        """Crea la hoja con el detalle de todas las pólizas"""
        # Título
        ws["A1"] = "DETALLE DE PÓLIZAS"
        ws["A1"].font = Font(size=14, bold=True)

        # Headers
        headers = [
            "Número de Póliza",
            "Compañía Aseguradora",
            "Corredor",
            "Tipo de Póliza",
            "Suma Asegurada",
            "Fecha Inicio",
            "Fecha Fin",
            "Días para Vencer",
            "Estado",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Datos
        row = 4
        for poliza in polizas:
            ws.cell(row=row, column=1, value=poliza.numero_poliza)
            ws.cell(row=row, column=2, value=str(poliza.compania_aseguradora))
            ws.cell(row=row, column=3, value=str(poliza.corredor_seguros))
            ws.cell(row=row, column=4, value=str(poliza.tipo_poliza))
            ws.cell(row=row, column=5, value=f"${poliza.suma_asegurada:,.2f}")
            ws.cell(row=row, column=6, value=poliza.fecha_inicio.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=7, value=poliza.fecha_fin.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=8, value=poliza.dias_para_vencer)
            ws.cell(row=row, column=9, value=poliza.get_estado_display())

            # Color según estado
            estado_cell = ws.cell(row=row, column=9)
            if poliza.estado == "vigente":
                estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif poliza.estado == "por_vencer":
                estado_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif poliza.estado == "vencida":
                estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def crear_hoja_gastos(self, ws, polizas):
        """Crea la hoja con los gastos por póliza - OPTIMIZADO usando datos prefetched"""
        # Título
        ws["A1"] = "GASTOS POR PÓLIZA"
        ws["A1"].font = Font(size=14, bold=True)

        # Headers
        headers = [
            "Número de Póliza",
            "Compañía",
            "Total Facturado",
            "Total Pagado",
            "Saldo Pendiente",
            "Número de Facturas",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Datos - OPTIMIZADO: aprovecha prefetch_related ya cargado
        row = 4
        total_facturado_global = Decimal("0")
        total_pagado_global = Decimal("0")

        for poliza in polizas:
            # Usar los datos ya prefetched (no genera consultas adicionales)
            facturas_list = list(poliza.facturas.all())
            total_facturado = sum((f.monto_total for f in facturas_list), Decimal("0"))

            # Los pagos ya están prefetched y filtrados por estado='aprobado'
            total_pagado = Decimal("0")
            for factura in facturas_list:
                # pagos ya está prefetched, no genera consultas
                pagos_list = list(factura.pagos.all())
                total_pagado += sum((p.monto for p in pagos_list), Decimal("0"))

            saldo_pendiente = total_facturado - total_pagado
            num_facturas = len(facturas_list)

            ws.cell(row=row, column=1, value=poliza.numero_poliza)
            ws.cell(row=row, column=2, value=str(poliza.compania_aseguradora))
            ws.cell(row=row, column=3, value=f"${total_facturado:,.2f}")
            ws.cell(row=row, column=4, value=f"${total_pagado:,.2f}")
            ws.cell(row=row, column=5, value=f"${saldo_pendiente:,.2f}")
            ws.cell(row=row, column=6, value=num_facturas)

            total_facturado_global += total_facturado
            total_pagado_global += total_pagado

            row += 1

        # Totales
        ws.cell(row=row, column=1, value="TOTALES")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"${total_facturado_global:,.2f}")
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"${total_pagado_global:,.2f}")
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"${total_facturado_global - total_pagado_global:,.2f}")
        ws.cell(row=row, column=5).font = Font(bold=True)

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def generar_pdf(self, polizas, filename, estado):
        """Genera el reporte en formato PDF"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Estilo personalizado para título
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#366092"),
            spaceAfter=30,
            alignment=1,  # Centrado
        )

        # Título
        title = Paragraph("REPORTE DE PÓLIZAS", title_style)
        elements.append(title)

        # Fecha
        fecha_text = f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        elements.append(Paragraph(fecha_text, styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Resumen por estado - OPTIMIZADO con una sola consulta
        elements.append(Paragraph("RESUMEN POR ESTADO", styles["Heading2"]))

        resumen_data = [["Estado", "Cantidad", "Suma Asegurada"]]

        # OPTIMIZACIÓN: una sola consulta para todos los estados
        resumen_estados = polizas.values("estado").annotate(
            cantidad=Count("id"), suma_total=Coalesce(Sum("suma_asegurada"), Decimal("0"))
        )
        estados_dict = {r["estado"]: r for r in resumen_estados}
        estados_orden = ["vigente", "vencida", "por_vencer", "cancelada"]

        for est in estados_orden:
            data = estados_dict.get(est, {"cantidad": 0, "suma_total": Decimal("0")})
            resumen_data.append([est.replace("_", " ").title(), str(data["cantidad"]), f"${data['suma_total']:,.2f}"])

        resumen_table = Table(resumen_data, colWidths=[2 * inch, 1.5 * inch, 2 * inch])
        resumen_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(resumen_table)
        elements.append(Spacer(1, 30))

        # Detalle de pólizas (primeras 20)
        elements.append(Paragraph("DETALLE DE PÓLIZAS", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        detalle_data = [["Número", "Compañía", "Suma Asegurada", "Vigencia", "Estado"]]

        for poliza in polizas[:20]:  # Limitar a 20 para el PDF
            detalle_data.append(
                [
                    poliza.numero_poliza[:15],
                    str(poliza.compania_aseguradora)[:20],
                    f"${poliza.suma_asegurada:,.0f}",
                    f'{poliza.fecha_inicio.strftime("%d/%m/%y")}\n{poliza.fecha_fin.strftime("%d/%m/%y")}',
                    poliza.get_estado_display(),
                ]
            )

        if polizas.count() > 20:
            detalle_data.append(["...", "...", "...", "...", "..."])

        detalle_table = Table(detalle_data, colWidths=[1.2 * inch, 1.8 * inch, 1.3 * inch, 1.2 * inch, 1 * inch])
        detalle_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(detalle_table)

        # Construir PDF
        doc.build(elements)

from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import Coalesce
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from app.models import Siniestro, TipoSiniestro, Poliza
import os


class Command(BaseCommand):
    help = 'Genera reportes de siniestros con análisis estadístico'

    def add_arguments(self, parser):
        parser.add_argument(
            '--formato',
            type=str,
            choices=['excel', 'pdf', 'ambos'],
            default='ambos',
            help='Formato del reporte: excel, pdf o ambos'
        )
        parser.add_argument(
            '--periodo',
            type=str,
            choices=['semanal', 'mensual', 'trimestral', 'anual', 'todo'],
            default='mensual',
            help='Período del reporte'
        )

    def handle(self, *args, **options):
        formato = options['formato']
        periodo = options['periodo']
        
        self.stdout.write(self.style.SUCCESS(f'Generando reporte de siniestros (período: {periodo})...'))
        
        # Obtener siniestros según período (optimizado con select_related)
        siniestros = self.obtener_siniestros_periodo(periodo)
        
        # Crear directorio de reportes si no existe
        reportes_dir = 'media/reportes/siniestros'
        os.makedirs(reportes_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        if formato in ['excel', 'ambos']:
            excel_file = f'{reportes_dir}/reporte_siniestros_{periodo}_{timestamp}.xlsx'
            self.generar_excel(siniestros, excel_file, periodo)
            self.stdout.write(self.style.SUCCESS(f'  ✓ Reporte Excel generado: {excel_file}'))
        
        if formato in ['pdf', 'ambos']:
            pdf_file = f'{reportes_dir}/reporte_siniestros_{periodo}_{timestamp}.pdf'
            self.generar_pdf(siniestros, pdf_file, periodo)
            self.stdout.write(self.style.SUCCESS(f'  ✓ Reporte PDF generado: {pdf_file}'))
        
        self.stdout.write(self.style.SUCCESS(f'✓ Reporte completado: {siniestros.count()} siniestros procesados'))

    def obtener_siniestros_periodo(self, periodo):
        """Obtiene los siniestros según el período especificado - OPTIMIZADO con select_related"""
        hoy = timezone.now()
        
        # Optimización: usar select_related para evitar consultas N+1
        base_queryset = Siniestro.objects.select_related(
            'poliza', 
            'poliza__compania_aseguradora', 
            'tipo_siniestro'
        )
        
        if periodo == 'semanal':
            fecha_inicio = hoy - timedelta(days=7)
        elif periodo == 'mensual':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestral':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:  # todo
            return base_queryset.all()
        
        return base_queryset.filter(fecha_siniestro__gte=fecha_inicio)

    def generar_excel(self, siniestros, filename, periodo):
        """Genera el reporte en formato Excel"""
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen Ejecutivo
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"
        self.crear_hoja_resumen(ws_resumen, siniestros, periodo)
        
        # Hoja 2: Detalle de Siniestros
        ws_detalle = wb.create_sheet("Detalle de Siniestros")
        self.crear_hoja_detalle(ws_detalle, siniestros)
        
        # Hoja 3: Análisis por Tipo
        ws_tipo = wb.create_sheet("Análisis por Tipo")
        self.crear_hoja_analisis_tipo(ws_tipo, siniestros)
        
        # Hoja 4: Análisis por Póliza
        ws_poliza = wb.create_sheet("Análisis por Póliza")
        self.crear_hoja_analisis_poliza(ws_poliza, siniestros)
        
        # Hoja 5: Tiempos de Resolución
        ws_tiempos = wb.create_sheet("Tiempos de Resolución")
        self.crear_hoja_tiempos(ws_tiempos, siniestros)
        
        wb.save(filename)

    def crear_hoja_resumen(self, ws, siniestros, periodo):
        """Crea la hoja de resumen ejecutivo - OPTIMIZADO con agregaciones SQL"""
        # Título
        ws['A1'] = f'REPORTE DE SINIESTROS - RESUMEN EJECUTIVO ({periodo.upper()})'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Estadísticas generales - OPTIMIZADO: una sola consulta con múltiples agregaciones
        ws['A4'] = 'ESTADÍSTICAS GENERALES'
        ws['A4'].font = Font(size=12, bold=True)
        
        # Agregación eficiente en una sola consulta
        stats = siniestros.aggregate(
            total=Count('id'),
            monto_total=Coalesce(Sum('monto_estimado'), Decimal('0')),
            monto_indemnizado=Coalesce(Sum('monto_indemnizado'), Decimal('0'))
        )
        total_siniestros = stats['total']
        monto_total = stats['monto_total']
        monto_indemnizado = stats['monto_indemnizado']
        tiempo_promedio = self.calcular_tiempo_promedio_resolucion(siniestros)
        
        estadisticas = [
            ['Total de Siniestros:', total_siniestros],
            ['Monto Total Estimado:', f'${monto_total:,.2f}'],
            ['Monto Total Indemnizado:', f'${monto_indemnizado:,.2f}'],
            ['Tiempo Promedio de Resolución:', f'{tiempo_promedio} días'],
        ]
        
        row = 5
        for label, valor in estadisticas:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=valor)
            row += 1
        
        # Resumen por estado - OPTIMIZADO: una sola consulta con annotate
        ws['A10'] = 'RESUMEN POR ESTADO'
        ws['A10'].font = Font(size=12, bold=True)
        
        headers = ['Estado', 'Cantidad', 'Monto Estimado', 'Porcentaje']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # OPTIMIZACIÓN: una sola consulta para todos los estados
        resumen_estados = siniestros.values('estado').annotate(
            cantidad=Count('id'),
            monto=Coalesce(Sum('monto_estimado'), Decimal('0'))
        ).order_by('estado')
        
        # Convertir a diccionario para acceso rápido
        estados_dict = {r['estado']: r for r in resumen_estados}
        
        estados_orden = [
            'registrado', 'documentacion_pendiente', 'enviado_aseguradora',
            'en_evaluacion', 'aprobado', 'rechazado', 'liquidado', 'cerrado'
        ]
        
        row = 12
        for estado in estados_orden:
            data = estados_dict.get(estado, {'cantidad': 0, 'monto': Decimal('0')})
            cantidad = data['cantidad']
            monto = data['monto']
            porcentaje = (cantidad / total_siniestros * 100) if total_siniestros > 0 else 0
            
            ws.cell(row=row, column=1, value=estado.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=cantidad)
            ws.cell(row=row, column=3, value=f'${monto:,.2f}')
            ws.cell(row=row, column=4, value=f'{porcentaje:.1f}%')
            row += 1
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

    def crear_hoja_detalle(self, ws, siniestros):
        """Crea la hoja con el detalle de todos los siniestros"""
        ws['A1'] = 'DETALLE DE SINIESTROS'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = [
            'Número', 'Tipo', 'Bien', 'Póliza', 'Fecha Siniestro',
            'Monto Estimado', 'Monto Indemnizado', 'Estado', 'Días desde Registro'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row = 4
        for siniestro in siniestros:
            ws.cell(row=row, column=1, value=siniestro.numero_siniestro)
            ws.cell(row=row, column=2, value=str(siniestro.tipo_siniestro))
            ws.cell(row=row, column=3, value=siniestro.bien_nombre)
            ws.cell(row=row, column=4, value=siniestro.poliza.numero_poliza)
            ws.cell(row=row, column=5, value=siniestro.fecha_siniestro.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=6, value=f'${siniestro.monto_estimado:,.2f}')
            ws.cell(row=row, column=7, value=f'${siniestro.monto_indemnizado:,.2f}' if siniestro.monto_indemnizado else 'N/A')
            ws.cell(row=row, column=8, value=siniestro.get_estado_display())
            ws.cell(row=row, column=9, value=siniestro.dias_desde_registro)
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def crear_hoja_analisis_tipo(self, ws, siniestros):
        """Crea la hoja con análisis por tipo de siniestro - OPTIMIZADO con agregaciones SQL"""
        ws['A1'] = 'ANÁLISIS POR TIPO DE SINIESTRO'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Tipo de Siniestro', 'Cantidad', 'Porcentaje', 'Monto Total', 'Monto Promedio']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        total_siniestros = siniestros.count()
        
        # OPTIMIZACIÓN: una sola consulta con aggregaciones en lugar de iterar TipoSiniestro
        stats_por_tipo = siniestros.values(
            'tipo_siniestro__nombre'
        ).annotate(
            cantidad=Count('id'),
            monto_total=Coalesce(Sum('monto_estimado'), Decimal('0')),
            monto_promedio=Coalesce(Avg('monto_estimado'), Decimal('0'))
        ).filter(cantidad__gt=0).order_by('-cantidad')
        
        row = 4
        for stat in stats_por_tipo:
            cantidad = stat['cantidad']
            porcentaje = (cantidad / total_siniestros * 100) if total_siniestros > 0 else 0
            
            ws.cell(row=row, column=1, value=stat['tipo_siniestro__nombre'] or 'Sin Tipo')
            ws.cell(row=row, column=2, value=cantidad)
            ws.cell(row=row, column=3, value=f'{porcentaje:.1f}%')
            ws.cell(row=row, column=4, value=f"${stat['monto_total']:,.2f}")
            ws.cell(row=row, column=5, value=f"${stat['monto_promedio']:,.2f}")
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def crear_hoja_analisis_poliza(self, ws, siniestros):
        """Crea la hoja con análisis por póliza - OPTIMIZADO con agregaciones SQL"""
        ws['A1'] = 'ANÁLISIS POR PÓLIZA (TOP 20)'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Número de Póliza', 'Compañía', 'Cantidad Siniestros', 'Monto Total']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # OPTIMIZACIÓN: agregación SQL directa en lugar de iterar en Python
        poliza_stats = siniestros.values(
            'poliza__numero_poliza',
            'poliza__compania_aseguradora__nombre'
        ).annotate(
            cantidad=Count('id'),
            monto_total=Coalesce(Sum('monto_estimado'), Decimal('0'))
        ).order_by('-cantidad')[:20]
        
        row = 4
        for stat in poliza_stats:
            ws.cell(row=row, column=1, value=stat['poliza__numero_poliza'])
            ws.cell(row=row, column=2, value=stat['poliza__compania_aseguradora__nombre'] or 'Sin Compañía')
            ws.cell(row=row, column=3, value=stat['cantidad'])
            ws.cell(row=row, column=4, value=f"${stat['monto_total']:,.2f}")
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def crear_hoja_tiempos(self, ws, siniestros):
        """Crea la hoja con análisis de tiempos de resolución"""
        ws['A1'] = 'ANÁLISIS DE TIEMPOS DE RESOLUCIÓN'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Siniestros cerrados
        siniestros_cerrados = siniestros.filter(estado='cerrado')
        
        ws['A3'] = f'Total de Siniestros Cerrados: {siniestros_cerrados.count()}'
        ws['A3'].font = Font(bold=True)
        
        headers = [
            'Número', 'Tipo', 'Fecha Registro', 'Fecha Liquidación',
            'Días de Gestión', 'Estado Final'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        row = 6
        tiempos = []
        
        for siniestro in siniestros_cerrados:
            if siniestro.fecha_liquidacion:
                dias_gestion = (siniestro.fecha_liquidacion - siniestro.fecha_registro.date()).days
                tiempos.append(dias_gestion)
                
                ws.cell(row=row, column=1, value=siniestro.numero_siniestro)
                ws.cell(row=row, column=2, value=str(siniestro.tipo_siniestro))
                ws.cell(row=row, column=3, value=siniestro.fecha_registro.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=4, value=siniestro.fecha_liquidacion.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=5, value=dias_gestion)
                ws.cell(row=row, column=6, value=siniestro.get_estado_display())
                row += 1
        
        # Estadísticas de tiempo
        if tiempos:
            ws[f'A{row + 2}'] = 'ESTADÍSTICAS DE TIEMPO'
            ws[f'A{row + 2}'].font = Font(bold=True)
            
            ws[f'A{row + 3}'] = f'Tiempo Mínimo: {min(tiempos)} días'
            ws[f'A{row + 4}'] = f'Tiempo Máximo: {max(tiempos)} días'
            ws[f'A{row + 5}'] = f'Tiempo Promedio: {sum(tiempos) / len(tiempos):.1f} días'
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def calcular_tiempo_promedio_resolucion(self, siniestros):
        """Calcula el tiempo promedio de resolución de siniestros"""
        siniestros_cerrados = siniestros.filter(
            estado='cerrado',
            fecha_liquidacion__isnull=False
        )
        
        if not siniestros_cerrados:
            return 0
        
        tiempos = []
        for siniestro in siniestros_cerrados:
            dias = (siniestro.fecha_liquidacion - siniestro.fecha_registro.date()).days
            tiempos.append(dias)
        
        return sum(tiempos) / len(tiempos) if tiempos else 0

    def generar_pdf(self, siniestros, filename, periodo):
        """Genera el reporte en formato PDF"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        
        # Título
        title = Paragraph(f'REPORTE DE SINIESTROS ({periodo.upper()})', title_style)
        elements.append(title)
        
        fecha_text = f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        elements.append(Paragraph(fecha_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Estadísticas generales
        total_siniestros = siniestros.count()
        monto_total = siniestros.aggregate(Sum('monto_estimado'))['monto_estimado__sum'] or 0
        tiempo_promedio = self.calcular_tiempo_promedio_resolucion(siniestros)
        
        elements.append(Paragraph('ESTADÍSTICAS GENERALES', styles['Heading2']))
        
        stats_data = [
            ['Total de Siniestros:', str(total_siniestros)],
            ['Monto Total Estimado:', f'${monto_total:,.2f}'],
            ['Tiempo Promedio de Resolución:', f'{tiempo_promedio:.1f} días'],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))
        
        # Resumen por tipo
        elements.append(Paragraph('ANÁLISIS POR TIPO DE SINIESTRO', styles['Heading2']))
        
        tipo_data = [['Tipo', 'Cantidad', 'Monto Total']]
        tipos = TipoSiniestro.objects.all()
        
        for tipo in tipos:
            siniestros_tipo = siniestros.filter(tipo_siniestro=tipo)
            cantidad = siniestros_tipo.count()
            if cantidad > 0:
                monto = siniestros_tipo.aggregate(Sum('monto_estimado'))['monto_estimado__sum'] or 0
                tipo_data.append([
                    str(tipo),
                    str(cantidad),
                    f'${monto:,.2f}'
                ])
        
        tipo_table = Table(tipo_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
        tipo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(tipo_table)
        
        doc.build(elements)

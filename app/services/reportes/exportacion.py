import csv

import io

from datetime import datetime

from decimal import Decimal

from django.http import HttpResponse

from django.utils import timezone

import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.utils import get_column_letter


def make_naive(dt):

    """

    Convierte un datetime con timezone a naive (sin timezone) para Excel.

    Excel no soporta datetimes con timezone.

    """

    if dt is None:

        return None

    if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:

        # Convertir a hora local y remover timezone

        return timezone.localtime(dt).replace(tzinfo=None)

    return dt



class ExportacionService:

    HEADER_FILL = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')

    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    THIN_BORDER = Border(

        left=Side(style='thin', color='E2E8F0'),

        right=Side(style='thin', color='E2E8F0'),

        top=Side(style='thin', color='E2E8F0'),

        bottom=Side(style='thin', color='E2E8F0')

    )

    @classmethod
    def exportar_polizas_csv(cls, queryset):

        response = HttpResponse(content_type='text/csv; charset=utf-8')

        response['Content-Disposition'] = f'attachment; filename="polizas_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

        response.write('\ufeff')

        writer = csv.writer(response)

        writer.writerow([

            'Número de Póliza', 'Compañía Aseguradora', 'Corredor', 'Tipo',

            'Suma Asegurada', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Días para Vencer'

        ])

        for poliza in queryset:

            writer.writerow([

                poliza.numero_poliza,

                str(poliza.compania_aseguradora),

                str(poliza.corredor_seguros),

                str(poliza.tipo_poliza),

                f'{poliza.suma_asegurada:.2f}',

                poliza.fecha_inicio.strftime('%d/%m/%Y'),

                poliza.fecha_fin.strftime('%d/%m/%Y'),

                poliza.get_estado_display(),

                poliza.dias_para_vencer

            ])

        return response

    @classmethod
    def exportar_polizas_excel(cls, queryset):

        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "Pólizas"

        headers = [

            'Número de Póliza', 'Compañía Aseguradora', 'Corredor', 'Tipo',

            'Suma Asegurada', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Días para Vencer'

        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col, value=header)

            cell.fill = cls.HEADER_FILL

            cell.font = cls.HEADER_FONT

            cell.alignment = cls.HEADER_ALIGNMENT

            cell.border = cls.THIN_BORDER

        estado_colors = {

            'vigente': 'C6F6D5',

            'por_vencer': 'FEFCBF',

            'vencida': 'FED7D7',

            'cancelada': 'E2E8F0'

        }

        for row_num, poliza in enumerate(queryset, start=2):

            data = [

                poliza.numero_poliza,

                str(poliza.compania_aseguradora),

                str(poliza.corredor_seguros),

                str(poliza.tipo_poliza),

                poliza.suma_asegurada,

                poliza.fecha_inicio,

                poliza.fecha_fin,

                poliza.get_estado_display(),

                poliza.dias_para_vencer

            ]

            for col, value in enumerate(data, start=1):

                cell = ws.cell(row=row_num, column=col, value=value)

                cell.border = cls.THIN_BORDER

                if col == 5:

                    cell.number_format = '"$"#,##0.00'

                elif col in [6, 7]:

                    cell.number_format = 'DD/MM/YYYY'

                elif col == 8:

                    color = estado_colors.get(poliza.estado, 'FFFFFF')

                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        for col in range(1, len(headers) + 1):

            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = 'A2'

        response = HttpResponse(

            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        )

        response['Content-Disposition'] = f'attachment; filename="polizas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        wb.save(response)

        return response

    @classmethod
    def exportar_facturas_csv(cls, queryset):

        response = HttpResponse(content_type='text/csv; charset=utf-8')

        response['Content-Disposition'] = f'attachment; filename="facturas_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

        response.write('\ufeff')

        writer = csv.writer(response)

        writer.writerow([

            'Número de Factura', 'Póliza', 'Compañía', 'Fecha Emisión',

            'Fecha Vencimiento', 'Subtotal', 'IVA', 'Contribuciones', 'Total', 'Estado'

        ])

        for factura in queryset:

            contribuciones = factura.contribucion_superintendencia + factura.contribucion_seguro_campesino

            writer.writerow([

                factura.numero_factura,

                factura.poliza.numero_poliza,

                str(factura.poliza.compania_aseguradora),

                factura.fecha_emision.strftime('%d/%m/%Y'),

                factura.fecha_vencimiento.strftime('%d/%m/%Y'),

                f'{factura.subtotal:.2f}',

                f'{factura.iva:.2f}',

                f'{contribuciones:.2f}',

                f'{factura.monto_total:.2f}',

                factura.get_estado_display()

            ])

        return response

    @classmethod
    def exportar_facturas_excel(cls, queryset):

        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "Facturas"

        headers = [

            'Número de Factura', 'Póliza', 'Compañía', 'Fecha Emisión',

            'Fecha Vencimiento', 'Subtotal', 'IVA', 'Contribuciones', 'Total', 'Saldo', 'Estado'

        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col, value=header)

            cell.fill = cls.HEADER_FILL

            cell.font = cls.HEADER_FONT

            cell.alignment = cls.HEADER_ALIGNMENT

            cell.border = cls.THIN_BORDER

        estado_colors = {

            'pendiente': 'FEFCBF',

            'pagada': 'C6F6D5',

            'parcial': 'BEE3F8',

            'vencida': 'FED7D7'

        }

        for row_num, factura in enumerate(queryset, start=2):

            contribuciones = factura.contribucion_superintendencia + factura.contribucion_seguro_campesino

            data = [

                factura.numero_factura,

                factura.poliza.numero_poliza,

                str(factura.poliza.compania_aseguradora),

                factura.fecha_emision,

                factura.fecha_vencimiento,

                factura.subtotal,

                factura.iva,

                contribuciones,

                factura.monto_total,

                factura.saldo_pendiente,

                factura.get_estado_display()

            ]

            for col, value in enumerate(data, start=1):

                cell = ws.cell(row=row_num, column=col, value=value)

                cell.border = cls.THIN_BORDER

                if col in [6, 7, 8, 9, 10]:

                    cell.number_format = '"$"#,##0.00'

                elif col in [4, 5]:

                    cell.number_format = 'DD/MM/YYYY'

                elif col == 11:

                    color = estado_colors.get(factura.estado, 'FFFFFF')

                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        for col in range(1, len(headers) + 1):

            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = 'A2'

        response = HttpResponse(

            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        )

        response['Content-Disposition'] = f'attachment; filename="facturas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        wb.save(response)

        return response

    @classmethod
    def exportar_siniestros_csv(cls, queryset):

        response = HttpResponse(content_type='text/csv; charset=utf-8')

        response['Content-Disposition'] = f'attachment; filename="siniestros_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

        response.write('\ufeff')

        writer = csv.writer(response)

        writer.writerow([

            'Número', 'Póliza', 'Tipo', 'Fecha', 'Bien', 'Ubicación',

            'Monto Estimado', 'Monto Indemnizado', 'Estado', 'Días en Gestión'

        ])

        for siniestro in queryset:

            writer.writerow([

                siniestro.numero_siniestro,

                siniestro.poliza.numero_poliza,

                str(siniestro.tipo_siniestro),

                siniestro.fecha_siniestro.strftime('%d/%m/%Y %H:%M'),

                siniestro.bien_nombre,

                siniestro.ubicacion,

                f'{siniestro.monto_estimado:.2f}',

                f'{siniestro.monto_indemnizado:.2f}' if siniestro.monto_indemnizado else 'N/A',

                siniestro.get_estado_display(),

                siniestro.dias_desde_registro

            ])

        return response

    @classmethod
    def exportar_siniestros_excel(cls, queryset):

        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "Siniestros"

        headers = [

            'Número', 'Póliza', 'Compañía', 'Tipo', 'Fecha', 'Bien',

            'Monto Estimado', 'Monto Indemnizado', 'Estado', 'Días en Gestión'

        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col, value=header)

            cell.fill = cls.HEADER_FILL

            cell.font = cls.HEADER_FONT

            cell.alignment = cls.HEADER_ALIGNMENT

            cell.border = cls.THIN_BORDER

        estado_colors = {

            'registrado': 'BEE3F8',

            'documentacion_pendiente': 'FEFCBF',

            'enviado_aseguradora': 'E9D8FD',

            'en_evaluacion': 'B2F5EA',

            'aprobado': 'C6F6D5',

            'rechazado': 'FED7D7',

            'liquidado': '9AE6B4',

            'cerrado': 'E2E8F0'

        }

        for row_num, siniestro in enumerate(queryset, start=2):

            data = [

                siniestro.numero_siniestro,

                siniestro.poliza.numero_poliza,

                str(siniestro.poliza.compania_aseguradora),

                str(siniestro.tipo_siniestro),

                make_naive(siniestro.fecha_siniestro),  # Remover timezone para Excel

                siniestro.bien_nombre,

                siniestro.monto_estimado,

                siniestro.monto_indemnizado or 0,

                siniestro.get_estado_display(),

                siniestro.dias_desde_registro

            ]

            for col, value in enumerate(data, start=1):

                cell = ws.cell(row=row_num, column=col, value=value)

                cell.border = cls.THIN_BORDER

                if col in [7, 8]:

                    cell.number_format = '"$"#,##0.00'

                elif col == 5:

                    cell.number_format = 'DD/MM/YYYY HH:MM'

                elif col == 9:

                    color = estado_colors.get(siniestro.estado, 'FFFFFF')

                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        for col in range(1, len(headers) + 1):

            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = 'A2'

        response = HttpResponse(

            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        )

        response['Content-Disposition'] = f'attachment; filename="siniestros_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        wb.save(response)

        return response

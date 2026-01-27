import mimetypes
import os
from decimal import Decimal
from itertools import chain

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView as DjangoLoginView
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView

from .forms import SubtipoRamoForm  # Alias de compatibilidad
from .forms import (  # Formularios de entidades base
    AdjuntoSiniestroFormSet,
    BienAseguradoForm,
    ChecklistSiniestroConfigForm,
    CompaniaAseguradoraForm,
    ConfiguracionBulkForm,
    ConfiguracionSistemaForm,
    CorredorSegurosForm,
    DetallePolizaRamoFormSet,
    DocumentoForm,
    FacturaForm,
    FiltroPolizasForm,
    FiltroReportesForm,
    FiltroSiniestrosForm,
    GrupoBienesForm,
    GrupoRamoForm,
    NotaCreditoForm,
    PagoForm,
    PolizaForm,
    RamoForm,
    ResponsableCustodioForm,
    SiniestroForm,
    SiniestroGestionForm,
    SubgrupoRamoForm,
    TipoRamoForm,
    TipoSiniestroForm,
)
from .models import SubtipoRamo  # Alias de compatibilidad
from .models import (
    AdjuntoSiniestro,
    Alerta,
    BackupRegistro,
    BienAsegurado,
    CalendarEvent,
    ChecklistSiniestro,
    ChecklistSiniestroConfig,
    CompaniaAseguradora,
    ConfiguracionBackup,
    ConfiguracionSistema,
    CorredorSeguros,
    DetallePolizaRamo,
    Documento,
    Factura,
    GrupoBienes,
    GrupoRamo,
    NotaCredito,
    NotificacionEmail,
    Pago,
    PaymentApproval,
    PolicyRenewal,
    Poliza,
    Quote,
    QuoteOption,
    Ramo,
    ResponsableCustodio,
    Siniestro,
    SiniestroEmail,
    SubgrupoRamo,
    TipoPoliza,
    TipoRamo,
    TipoSiniestro,
)
from .services import (
    AdvancedAnalyticsService,
    DashboardAnalyticsService,
    DashboardFiltersService,
    DateRangePresets,
    DocumentosService,
    EstadisticasService,
    ExportacionService,
    NotificacionesService,
    ReportesAvanzadosService,
    ReportesService,
)
from .services.reportes import PDFReportesService


@login_required
def dashboard(request):
    """
    Vista principal del dashboard con filtros avanzados estilo Odoo.
    Soporta date range picker, filtros por entidad, y actualización dinámica.
    """
    import json

    # Parsear filtros desde la request
    filters = DashboardFiltersService.parse_filters_from_request(request)

    # Obtener filtros disponibles para los selectores
    available_filters = DashboardFiltersService.get_available_filters()

    # Obtener estadísticas filtradas
    filtered_stats = DashboardFiltersService.get_filtered_stats(filters)

    # Obtener datos para gráficos filtrados
    chart_data = DashboardFiltersService.get_chart_data(filters)

    # Obtener listas de registros
    lists_data = DashboardFiltersService.get_lists_data(filters)

    # Estadísticas globales (sin filtros) para KPIs generales
    kpis = EstadisticasService.get_kpis()

    # Datos comparativos y tendencias (usando el sistema anterior para compatibilidad)
    period_type = request.GET.get("period", DashboardAnalyticsService.PERIOD_MONTH)
    if period_type not in DashboardAnalyticsService.PERIOD_CHOICES:
        period_type = DashboardAnalyticsService.PERIOD_MONTH

    comparative_data = DashboardAnalyticsService.get_comparative_stats(period_type)
    year_comparison = DashboardAnalyticsService.get_year_over_year_comparison()
    quick_actions = DashboardAnalyticsService.get_quick_actions_data()

    context = {
        # Filtros
        "filters": filters,
        "available_filters": available_filters,
        "date_presets": DateRangePresets.CHOICES,
        # Estadísticas filtradas
        "stats": filtered_stats,
        "chart_data": json.dumps(chart_data),
        # Listas de registros
        "expiring_policies": lists_data["expiring_policies"],
        "pending_invoices": lists_data["pending_invoices"],
        "active_claims": lists_data["active_claims"],
        # KPIs globales
        "kpis": kpis,
        # Datos comparativos
        "period_type": period_type,
        "period_choices": DashboardAnalyticsService.PERIOD_CHOICES,
        "comparative": comparative_data,
        "year_comparison": json.dumps(year_comparison),
        "quick_actions": quick_actions,
    }

    return render(request, "app/dashboard/index.html", context)


@login_required
def polizas_lista(request):
    polizas = Poliza.objects.select_related("compania_aseguradora", "corredor_seguros", "tipo_poliza").order_by(
        "-fecha_inicio"
    )

    query = request.GET.get("q", "").strip()
    if query:
        polizas = polizas.filter(
            Q(numero_poliza__icontains=query)
            | Q(compania_aseguradora__nombre__icontains=query)
            | Q(corredor_seguros__nombre__icontains=query)
        )

    estado = request.GET.get("estado")
    if estado:
        polizas = polizas.filter(estado=estado)

    compania = request.GET.get("compania")
    if compania:
        polizas = polizas.filter(compania_aseguradora_id=compania)

    paginator = Paginator(polizas, 15)
    page = request.GET.get("page", 1)
    polizas_page = paginator.get_page(page)

    companias = CompaniaAseguradora.objects.filter(activo=True).order_by("nombre")

    context = {
        "polizas": polizas_page,
        "estado_filtro": estado,
        "query": query,
        "companias": companias,
        "compania_filtro": compania,
        "total_resultados": paginator.count,
    }

    return render(request, "app/polizas/lista.html", context)


@login_required
@require_GET
def polizas_exportar(request):
    formato = request.GET.get("formato", "excel")

    polizas = Poliza.objects.select_related("compania_aseguradora", "corredor_seguros", "tipo_poliza")

    estado = request.GET.get("estado")
    if estado:
        polizas = polizas.filter(estado=estado)

    if formato == "csv":
        return ExportacionService.exportar_polizas_csv(polizas)
    return ExportacionService.exportar_polizas_excel(polizas)


@login_required
def desglose_ramos_lista(request):
    """Vista para el módulo de Desglose por Ramos con filtros avanzados"""
    from django.db.models import Sum

    detalles = DetallePolizaRamo.objects.select_related(
        "poliza", "poliza__compania_aseguradora", "poliza__grupo_ramo", "subgrupo_ramo"
    ).order_by("-poliza__fecha_inicio", "subgrupo_ramo__nombre")

    # Filtro de búsqueda
    query = request.GET.get("q", "").strip()
    if query:
        detalles = detalles.filter(
            Q(poliza__facturas__numero_factura__icontains=query)
            | Q(poliza__facturas__documento_contable__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
        ).distinct()

    # Filtro por compañía
    compania = request.GET.get("compania")
    if compania:
        detalles = detalles.filter(poliza__compania_aseguradora_id=compania)

    # Filtro por grupo de ramo (muestra todos los subramos del grupo)
    ramo = request.GET.get("ramo")
    if ramo:
        detalles = detalles.filter(subgrupo_ramo__grupo_ramo_id=ramo)

    # Filtro por póliza
    poliza = request.GET.get("poliza")
    if poliza:
        detalles = detalles.filter(poliza_id=poliza)

    # Filtro por estado de póliza
    estado = request.GET.get("estado")
    if estado:
        detalles = detalles.filter(poliza__estado=estado)

    # Filtro por fechas (fecha inicio de póliza)
    fecha_desde = request.GET.get("fecha_desde")
    if fecha_desde:
        detalles = detalles.filter(poliza__fecha_inicio__gte=fecha_desde)

    fecha_hasta = request.GET.get("fecha_hasta")
    if fecha_hasta:
        detalles = detalles.filter(poliza__fecha_inicio__lte=fecha_hasta)

    # Calcular totales antes de paginar
    totales = detalles.aggregate(
        suma_asegurada=Sum("suma_asegurada"),
        prima=Sum("total_prima"),
        contribucion_super=Sum("contribucion_superintendencia"),
        emision=Sum("emision"),
        seguro_campesino=Sum("seguro_campesino"),
        base_imponible=Sum("base_imponible"),
        iva=Sum("iva"),
        facturado=Sum("total_facturado"),
        retencion_prima=Sum("retencion_prima"),
        retencion_iva=Sum("retencion_iva"),
        por_pagar=Sum("valor_por_pagar"),
    )
    totales["retenciones"] = (totales.get("retencion_prima") or 0) + (totales.get("retencion_iva") or 0)

    # Optimización: Pre-calcular total_retenciones con annotate() en lugar de loop
    from decimal import Decimal

    from django.db.models import DecimalField, F, Value
    from django.db.models.functions import Coalesce

    detalles = detalles.annotate(
        total_retenciones=Coalesce(F("retencion_prima"), Value(Decimal("0")), output_field=DecimalField())
        + Coalesce(F("retencion_iva"), Value(Decimal("0")), output_field=DecimalField())
    )

    # Paginación
    paginator = Paginator(detalles, 20)
    page = request.GET.get("page", 1)
    detalles_page = paginator.get_page(page)

    # Datos para filtros
    companias = CompaniaAseguradora.objects.filter(activo=True).order_by("nombre")
    ramos = GrupoRamo.objects.filter(activo=True).order_by("nombre")
    polizas_lista = Poliza.objects.order_by("-fecha_inicio")[:100]  # Limitar para rendimiento

    context = {
        "detalles": detalles_page,
        "totales": totales,
        "query": query,
        "companias": companias,
        "compania_filtro": compania,
        "ramos": ramos,
        "ramo_filtro": ramo,
        "polizas_lista": polizas_lista,
        "poliza_filtro": poliza,
        "estado_filtro": estado,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "total_resultados": paginator.count,
    }

    return render(request, "app/desglose_ramos/lista.html", context)


@login_required
@require_GET
def desglose_ramos_exportar(request):
    """Exportar desglose por ramos a Excel o CSV"""
    import csv

    from django.db.models import Sum

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    formato = request.GET.get("formato", "excel")

    detalles = DetallePolizaRamo.objects.select_related(
        "poliza", "poliza__compania_aseguradora", "poliza__grupo_ramo", "subgrupo_ramo"
    ).order_by("-poliza__fecha_inicio", "subgrupo_ramo__nombre")

    # Aplicar mismos filtros que la vista
    query = request.GET.get("q", "").strip()
    if query:
        detalles = detalles.filter(
            Q(poliza__facturas__numero_factura__icontains=query)
            | Q(poliza__facturas__documento_contable__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
        ).distinct()

    compania = request.GET.get("compania")
    if compania:
        detalles = detalles.filter(poliza__compania_aseguradora_id=compania)

    ramo = request.GET.get("ramo")
    if ramo:
        detalles = detalles.filter(subgrupo_ramo__grupo_ramo_id=ramo)

    poliza = request.GET.get("poliza")
    if poliza:
        detalles = detalles.filter(poliza_id=poliza)

    estado = request.GET.get("estado")
    if estado:
        detalles = detalles.filter(poliza__estado=estado)

    fecha_desde = request.GET.get("fecha_desde")
    if fecha_desde:
        detalles = detalles.filter(poliza__fecha_inicio__gte=fecha_desde)

    fecha_hasta = request.GET.get("fecha_hasta")
    if fecha_hasta:
        detalles = detalles.filter(poliza__fecha_inicio__lte=fecha_hasta)

    if formato == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="desglose_ramos.csv"'
        response.write("\ufeff")  # BOM para Excel

        writer = csv.writer(response)
        writer.writerow(
            [
                "Ramo",
                "N° Factura",
                "Doc. Contable",
                "N° Póliza",
                "Compañía",
                "Suma Asegurada",
                "Total Prima",
                "Cont. Superintendencia",
                "Emisión",
                "Seg. Campesino",
                "Base Imponible",
                "IVA 15%",
                "Total Facturado",
                "Retención Prima",
                "Retención IVA",
                "Valor por Pagar",
            ]
        )

        for d in detalles:
            writer.writerow(
                [
                    d.subgrupo_ramo.nombre if d.subgrupo_ramo else "-",
                    d.numero_factura or "",
                    d.documento_contable or "",
                    d.poliza.numero_poliza,
                    d.poliza.compania_aseguradora.nombre,
                    float(d.suma_asegurada),
                    float(d.total_prima),
                    float(d.contribucion_superintendencia),
                    float(d.emision),
                    float(d.seguro_campesino),
                    float(d.base_imponible),
                    float(d.iva),
                    float(d.total_facturado),
                    float(d.retencion_prima),
                    float(d.retencion_iva),
                    float(d.valor_por_pagar),
                ]
            )

        return response

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Desglose por Ramos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Encabezados
    headers = [
        "Ramo",
        "N° Factura",
        "Doc. Contable",
        "N° Póliza",
        "Compañía",
        "Suma Asegurada",
        "Total Prima",
        "Cont. Superintendencia",
        "Emisión",
        "Seg. Campesino",
        "Base Imponible",
        "IVA 15%",
        "Total Facturado",
        "Retención Prima",
        "Retención IVA",
        "Valor por Pagar",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row, d in enumerate(detalles, 2):
        ws.cell(row=row, column=1, value=d.subgrupo_ramo.nombre if d.subgrupo_ramo else "-").border = border
        ws.cell(row=row, column=2, value=d.numero_factura or "").border = border
        ws.cell(row=row, column=3, value=d.documento_contable or "").border = border
        ws.cell(row=row, column=4, value=d.poliza.numero_poliza).border = border
        ws.cell(row=row, column=5, value=d.poliza.compania_aseguradora.nombre).border = border
        ws.cell(row=row, column=6, value=float(d.suma_asegurada)).border = border
        ws.cell(row=row, column=7, value=float(d.total_prima)).border = border
        ws.cell(row=row, column=8, value=float(d.contribucion_superintendencia)).border = border
        ws.cell(row=row, column=9, value=float(d.emision)).border = border
        ws.cell(row=row, column=10, value=float(d.seguro_campesino)).border = border
        ws.cell(row=row, column=11, value=float(d.base_imponible)).border = border
        ws.cell(row=row, column=12, value=float(d.iva)).border = border
        ws.cell(row=row, column=13, value=float(d.total_facturado)).border = border
        ws.cell(row=row, column=14, value=float(d.retencion_prima)).border = border
        ws.cell(row=row, column=15, value=float(d.retencion_iva)).border = border
        ws.cell(row=row, column=16, value=float(d.valor_por_pagar)).border = border

        # Formato numérico
        for col in range(6, 17):
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    # Ajustar anchos
    column_widths = [20, 12, 15, 12, 25, 15, 12, 18, 10, 14, 15, 12, 15, 14, 14, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="desglose_ramos.xlsx"'
    wb.save(response)

    return response


@login_required
def facturas_lista(request):
    facturas = Factura.objects.select_related("poliza", "poliza__compania_aseguradora").order_by("-fecha_emision")

    query = request.GET.get("q", "").strip()
    if query:
        facturas = facturas.filter(
            Q(numero_factura__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
            | Q(poliza__compania_aseguradora__nombre__icontains=query)
        )

    estado = request.GET.get("estado")
    if estado:
        facturas = facturas.filter(estado=estado)

    paginator = Paginator(facturas, 15)
    page = request.GET.get("page", 1)
    facturas_page = paginator.get_page(page)

    context = {
        "facturas": facturas_page,
        "estado_filtro": estado,
        "query": query,
        "total_resultados": paginator.count,
    }

    return render(request, "app/facturas/lista.html", context)


@login_required
@require_GET
def facturas_exportar(request):
    formato = request.GET.get("formato", "excel")

    facturas = Factura.objects.select_related("poliza", "poliza__compania_aseguradora")

    estado = request.GET.get("estado")
    if estado:
        facturas = facturas.filter(estado=estado)

    if formato == "csv":
        return ExportacionService.exportar_facturas_csv(facturas)
    return ExportacionService.exportar_facturas_excel(facturas)


@login_required
def siniestros_lista(request):
    siniestros = Siniestro.objects.select_related("poliza", "tipo_siniestro", "poliza__compania_aseguradora").order_by(
        "-fecha_siniestro"
    )

    query = request.GET.get("q", "").strip()
    if query:
        siniestros = siniestros.filter(
            Q(numero_siniestro__icontains=query)
            | Q(bien_nombre__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
        )

    estado = request.GET.get("estado")
    if estado:
        siniestros = siniestros.filter(estado=estado)

    tipo = request.GET.get("tipo")
    if tipo:
        siniestros = siniestros.filter(tipo_siniestro_id=tipo)

    paginator = Paginator(siniestros, 15)
    page = request.GET.get("page", 1)
    siniestros_page = paginator.get_page(page)

    tipos = TipoSiniestro.objects.filter(activo=True)

    context = {
        "siniestros": siniestros_page,
        "estado_filtro": estado,
        "query": query,
        "tipos": tipos,
        "tipo_filtro": tipo,
        "total_resultados": paginator.count,
    }

    return render(request, "app/siniestros/lista.html", context)


@login_required
@require_GET
def siniestros_exportar(request):
    formato = request.GET.get("formato", "excel")

    siniestros = Siniestro.objects.select_related("poliza", "tipo_siniestro", "poliza__compania_aseguradora")

    estado = request.GET.get("estado")
    if estado:
        siniestros = siniestros.filter(estado=estado)

    if formato == "csv":
        return ExportacionService.exportar_siniestros_csv(siniestros)
    return ExportacionService.exportar_siniestros_excel(siniestros)


@login_required
def reportes_dashboard(request):
    stats = EstadisticasService.get_dashboard_stats()
    kpis = EstadisticasService.get_kpis()

    graficos_polizas = ReportesService.get_datos_graficos_polizas()
    graficos_facturas = ReportesService.get_datos_graficos_facturas()
    graficos_siniestros = ReportesService.get_datos_graficos_siniestros_mensual()

    polizas_por_compania = EstadisticasService.get_polizas_por_compania()
    polizas_por_tipo = EstadisticasService.get_polizas_por_tipo()
    siniestros_por_tipo = EstadisticasService.get_siniestros_por_tipo()

    context = {
        "stats": stats,
        "kpis": kpis,
        "graficos_polizas": graficos_polizas,
        "graficos_facturas": graficos_facturas,
        "graficos_siniestros": graficos_siniestros,
        "polizas_por_compania": polizas_por_compania,
        "polizas_por_tipo": polizas_por_tipo,
        "siniestros_por_tipo": siniestros_por_tipo,
    }

    return render(request, "app/reportes/dashboard.html", context)


@login_required
def reportes_polizas(request):
    filtros = {
        "estado": request.GET.get("estado"),
        "compania": request.GET.get("compania"),
        "tipo": request.GET.get("tipo"),
        "fecha_desde": request.GET.get("fecha_desde"),
        "fecha_hasta": request.GET.get("fecha_hasta"),
    }

    filtros = {k: v for k, v in filtros.items() if v}

    reporte = ReportesService.generar_reporte_polizas(filtros)

    paginator = Paginator(reporte["queryset"], 20)
    page = request.GET.get("page", 1)
    polizas_page = paginator.get_page(page)

    companias = CompaniaAseguradora.objects.filter(activo=True)
    tipos = TipoPoliza.objects.filter(activo=True)

    context = {
        "polizas": polizas_page,
        "totales": reporte["totales"],
        "por_compania": reporte["por_compania"],
        "por_tipo": reporte["por_tipo"],
        "companias": companias,
        "tipos": tipos,
        "filtros": filtros,
    }

    return render(request, "app/reportes/polizas.html", context)


@login_required
def reportes_siniestros(request):
    filtros = {
        "estado": request.GET.get("estado"),
        "tipo": request.GET.get("tipo"),
        "fecha_desde": request.GET.get("fecha_desde"),
        "fecha_hasta": request.GET.get("fecha_hasta"),
    }

    filtros = {k: v for k, v in filtros.items() if v}

    reporte = ReportesService.generar_reporte_siniestros(filtros)

    paginator = Paginator(reporte["queryset"], 20)
    page = request.GET.get("page", 1)
    siniestros_page = paginator.get_page(page)

    tipos = TipoSiniestro.objects.filter(activo=True)

    context = {
        "siniestros": siniestros_page,
        "totales": reporte["totales"],
        "por_tipo": reporte["por_tipo"],
        "por_estado": reporte["por_estado"],
        "tipos": tipos,
        "filtros": filtros,
    }

    return render(request, "app/reportes/siniestros.html", context)


@login_required
def reportes_polizas_pdf(request):
    """Exporta el reporte de pólizas en formato PDF."""
    filtros = {
        "estado": request.GET.get("estado"),
        "compania": request.GET.get("compania"),
        "tipo": request.GET.get("tipo"),
        "fecha_desde": request.GET.get("fecha_desde"),
        "fecha_hasta": request.GET.get("fecha_hasta"),
    }

    filtros = {k: v for k, v in filtros.items() if v}

    # Construir texto de filtros para el reporte
    filtros_texto = []
    if filtros.get("estado"):
        filtros_texto.append(f"Estado: {filtros['estado']}")
    if filtros.get("compania"):
        try:
            compania = CompaniaAseguradora.objects.get(id=filtros["compania"])
            filtros_texto.append(f"Compañía: {compania.nombre}")
        except CompaniaAseguradora.DoesNotExist:
            pass
    if filtros.get("tipo"):
        try:
            tipo = TipoPoliza.objects.get(id=filtros["tipo"])
            filtros_texto.append(f"Tipo: {tipo.nombre}")
        except TipoPoliza.DoesNotExist:
            pass
    if filtros.get("fecha_desde"):
        filtros_texto.append(f"Desde: {filtros['fecha_desde']}")
    if filtros.get("fecha_hasta"):
        filtros_texto.append(f"Hasta: {filtros['fecha_hasta']}")

    reporte = ReportesService.generar_reporte_polizas(filtros)

    return PDFReportesService.generar_reporte_polizas_pdf(
        reporte, filtros_texto=" | ".join(filtros_texto) if filtros_texto else None
    )


@login_required
def reportes_siniestros_pdf(request):
    """Exporta el reporte de siniestros en formato PDF."""
    filtros = {
        "estado": request.GET.get("estado"),
        "tipo": request.GET.get("tipo"),
        "fecha_desde": request.GET.get("fecha_desde"),
        "fecha_hasta": request.GET.get("fecha_hasta"),
    }

    filtros = {k: v for k, v in filtros.items() if v}

    # Construir texto de filtros para el reporte
    filtros_texto = []
    if filtros.get("estado"):
        filtros_texto.append(f"Estado: {filtros['estado']}")
    if filtros.get("tipo"):
        try:
            tipo = TipoSiniestro.objects.get(id=filtros["tipo"])
            filtros_texto.append(f"Tipo: {tipo.nombre}")
        except TipoSiniestro.DoesNotExist:
            pass
    if filtros.get("fecha_desde"):
        filtros_texto.append(f"Desde: {filtros['fecha_desde']}")
    if filtros.get("fecha_hasta"):
        filtros_texto.append(f"Hasta: {filtros['fecha_hasta']}")

    reporte = ReportesService.generar_reporte_siniestros(filtros)

    return PDFReportesService.generar_reporte_siniestros_pdf(
        reporte, filtros_texto=" | ".join(filtros_texto) if filtros_texto else None
    )


@login_required
def reportes_facturas_pdf(request):
    """Exporta el reporte de facturas en formato PDF."""
    filtros = {
        "estado": request.GET.get("estado"),
        "fecha_desde": request.GET.get("fecha_desde"),
        "fecha_hasta": request.GET.get("fecha_hasta"),
    }

    filtros = {k: v for k, v in filtros.items() if v}

    # Construir texto de filtros para el reporte
    filtros_texto = []
    if filtros.get("estado"):
        filtros_texto.append(f"Estado: {filtros['estado']}")
    if filtros.get("fecha_desde"):
        filtros_texto.append(f"Desde: {filtros['fecha_desde']}")
    if filtros.get("fecha_hasta"):
        filtros_texto.append(f"Hasta: {filtros['fecha_hasta']}")

    reporte = ReportesService.generar_reporte_facturas(filtros)

    return PDFReportesService.generar_reporte_facturas_pdf(
        reporte, filtros_texto=" | ".join(filtros_texto) if filtros_texto else None
    )


@login_required
def reportes_ejecutivo_pdf(request):
    """Genera un reporte ejecutivo general en PDF."""
    stats = EstadisticasService.get_dashboard_stats()
    kpis = EstadisticasService.get_kpis()

    dashboard_data = {
        "stats": stats,
        "kpis": kpis,
    }

    return PDFReportesService.generar_reporte_ejecutivo_pdf(dashboard_data)


@login_required
def alertas_lista(request):
    alertas = (
        Alerta.objects.select_related("poliza", "factura", "siniestro")
        .prefetch_related("destinatarios")
        .order_by("-fecha_creacion")
    )

    estado = request.GET.get("estado")
    if estado:
        alertas = alertas.filter(estado=estado)

    tipo = request.GET.get("tipo")
    if tipo:
        alertas = alertas.filter(tipo_alerta=tipo)

    paginator = Paginator(alertas, 20)
    page = request.GET.get("page", 1)
    alertas_page = paginator.get_page(page)

    context = {
        "alertas": alertas_page,
        "estado_filtro": estado,
        "tipo_filtro": tipo,
    }

    return render(request, "app/alertas/lista.html", context)


@login_required
@require_POST
def alerta_marcar_leida(request, pk):
    alerta = get_object_or_404(Alerta, pk=pk)
    alerta.marcar_como_leida()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"status": "success"})
    return redirect("alertas_lista")


@login_required
def api_stats(request):
    stats = EstadisticasService.get_dashboard_stats()
    return JsonResponse(stats)


@login_required
def api_kpis(request):
    kpis = EstadisticasService.get_kpis()
    kpis_serializable = {
        k: float(v) if isinstance(v, (int, float)) or hasattr(v, "__float__") else v for k, v in kpis.items()
    }
    return JsonResponse(kpis_serializable)


# ============================================================================
# DOCUMENTOS
# ============================================================================


@login_required
def documentos_lista(request):
    documentos = Documento.objects.select_related("poliza", "siniestro", "factura", "subido_por").order_by(
        "-fecha_subida"
    )

    # Filtros
    tipo = request.GET.get("tipo")
    if tipo:
        documentos = documentos.filter(tipo_documento=tipo)

    query = request.GET.get("q")
    if query:
        documentos = documentos.filter(
            Q(nombre__icontains=query)
            | Q(descripcion__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
            | Q(siniestro__numero_siniestro__icontains=query)
            | Q(factura__numero_factura__icontains=query)
        )

    # Paginación
    paginator = Paginator(documentos, 20)
    page = request.GET.get("page")
    documentos = paginator.get_page(page)

    context = {
        "documentos": documentos,
        "query": query or "",
        "tipo_filtro": tipo or "",
        "tipos_documento": Documento.TIPO_DOCUMENTO_CHOICES,
        "total_resultados": paginator.count,
    }

    return render(request, "app/documentos/lista.html", context)


@login_required
def documento_ver(request, pk):
    documento = get_object_or_404(Documento, pk=pk)

    # Obtener información del archivo
    file_path = documento.archivo.path
    file_name = os.path.basename(documento.archivo.name)
    file_ext = os.path.splitext(file_name)[1].lower()

    # Determinar el tipo de contenido
    content_type, _ = mimetypes.guess_type(file_path)

    # Categorizar el tipo de archivo para el template
    if file_ext in [".pdf"]:
        file_type = "pdf"
    elif file_ext in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"]:
        file_type = "image"
    elif file_ext in [".doc", ".docx"]:
        file_type = "word"
    elif file_ext in [".xls", ".xlsx"]:
        file_type = "excel"
    elif file_ext in [".txt", ".csv"]:
        file_type = "text"
    else:
        file_type = "other"

    context = {
        "documento": documento,
        "file_type": file_type,
        "file_ext": file_ext,
        "content_type": content_type,
    }

    return render(request, "app/documentos/ver.html", context)


@login_required
def documento_descargar(request, pk):
    documento = get_object_or_404(Documento, pk=pk)

    file_path = documento.archivo.path
    file_name = os.path.basename(documento.archivo.name)
    content_type, _ = mimetypes.guess_type(file_path)

    response = FileResponse(open(file_path, "rb"), content_type=content_type or "application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    return response


# ============================================================================
# BÚSQUEDA GLOBAL
# ============================================================================


@login_required
def busqueda_global(request):
    query = request.GET.get("q", "").strip()

    polizas = []
    facturas = []
    siniestros = []

    if query:
        polizas = Poliza.objects.filter(
            Q(numero_poliza__icontains=query)
            | Q(coberturas__icontains=query)
            | Q(observaciones__icontains=query)
            | Q(compania_aseguradora__nombre__icontains=query)
            | Q(corredor_seguros__nombre__icontains=query)
        ).select_related("compania_aseguradora", "corredor_seguros", "tipo_poliza")[:20]

        facturas = Factura.objects.filter(
            Q(numero_factura__icontains=query) | Q(poliza__numero_poliza__icontains=query)
        ).select_related("poliza")[:20]

        siniestros = Siniestro.objects.filter(
            Q(numero_siniestro__icontains=query)
            | Q(descripcion_detallada__icontains=query)
            | Q(ubicacion__icontains=query)
            | Q(poliza__numero_poliza__icontains=query)
        ).select_related("poliza", "tipo_siniestro")[:20]

    context = {
        "query": query,
        "polizas": polizas,
        "facturas": facturas,
        "siniestros": siniestros,
        "total_resultados": len(polizas) + len(facturas) + len(siniestros),
    }

    return render(request, "app/busqueda/resultados.html", context)


@login_required
def api_buscar(request):
    query = request.GET.get("q", "").strip()
    results = []

    if len(query) >= 2:
        try:
            # Buscar pólizas
            polizas = Poliza.objects.filter(
                Q(numero_poliza__icontains=query)
                | Q(compania_aseguradora__nombre__icontains=query)
                | Q(corredor_seguros__nombre__icontains=query)
            ).select_related("tipo_poliza", "compania_aseguradora")[:5]

            for p in polizas:
                results.append(
                    {
                        "type": "poliza",
                        "title": p.numero_poliza,
                        "subtitle": f"{p.compania_aseguradora.nombre} • {p.get_estado_display()}",
                        "url": f"/admin/app/poliza/{p.pk}/change/",
                    }
                )

            # Buscar facturas
            facturas = Factura.objects.filter(
                Q(numero_factura__icontains=query) | Q(poliza__numero_poliza__icontains=query)
            ).select_related("poliza")[:5]

            for f in facturas:
                monto = float(f.monto_total) if f.monto_total else 0
                results.append(
                    {
                        "type": "factura",
                        "title": f.numero_factura,
                        "subtitle": f"${monto:,.2f} • {f.get_estado_display()}",
                        "url": f"/admin/app/factura/{f.pk}/change/",
                    }
                )

            # Buscar siniestros
            siniestros = Siniestro.objects.filter(
                Q(numero_siniestro__icontains=query)
                | Q(descripcion_detallada__icontains=query)
                | Q(poliza__numero_poliza__icontains=query)
            ).select_related("poliza")[:5]

            for s in siniestros:
                desc = (
                    s.descripcion_detallada[:40] + "..."
                    if s.descripcion_detallada and len(s.descripcion_detallada) > 40
                    else (s.descripcion_detallada or "Sin descripción")
                )
                results.append(
                    {
                        "type": "siniestro",
                        "title": s.numero_siniestro,
                        "subtitle": f"{s.get_estado_display()} • {desc}",
                        "url": f"/admin/app/siniestro/{s.pk}/change/",
                    }
                )

        except Exception as e:
            return JsonResponse({"results": [], "error": str(e)}, status=200)

    return JsonResponse({"results": results})


# ============================================================================
# DASHBOARD ANALYTICS API
# ============================================================================


@login_required
@require_GET
def api_dashboard_summary(request):
    """
    Endpoint para obtener un resumen completo del dashboard.
    Parámetros:
        - period: Tipo de período (year, month, week, day)
    """
    period_type = request.GET.get("period", DashboardAnalyticsService.PERIOD_MONTH)
    if period_type not in DashboardAnalyticsService.PERIOD_CHOICES:
        period_type = DashboardAnalyticsService.PERIOD_MONTH

    summary = DashboardAnalyticsService.get_dashboard_summary(period_type)
    return JsonResponse(summary)


@login_required
@require_GET
def api_dashboard_comparative(request):
    """
    Endpoint para estadísticas comparativas entre períodos.
    Parámetros:
        - period: Tipo de período (year, month, week, day)
    """
    period_type = request.GET.get("period", DashboardAnalyticsService.PERIOD_MONTH)
    if period_type not in DashboardAnalyticsService.PERIOD_CHOICES:
        period_type = DashboardAnalyticsService.PERIOD_MONTH

    data = DashboardAnalyticsService.get_comparative_stats(period_type)
    return JsonResponse(data)


@login_required
@require_GET
def api_dashboard_trend(request):
    """
    Endpoint para datos de tendencia histórica.
    Parámetros:
        - period: Tipo de período (year, month, week, day)
        - count: Cantidad de períodos (default: 12)
    """
    period_type = request.GET.get("period", DashboardAnalyticsService.PERIOD_MONTH)
    if period_type not in DashboardAnalyticsService.PERIOD_CHOICES:
        period_type = DashboardAnalyticsService.PERIOD_MONTH

    try:
        periods_count = int(request.GET.get("count", 12))
        periods_count = min(max(periods_count, 1), 24)  # Entre 1 y 24 períodos
    except ValueError:
        periods_count = 12

    data = DashboardAnalyticsService.get_trend_data(period_type, periods_count)
    return JsonResponse(data)


@login_required
@require_GET
def api_dashboard_year_comparison(request):
    """
    Endpoint para comparación año a año.
    """
    data = DashboardAnalyticsService.get_year_over_year_comparison()
    return JsonResponse(data)


@login_required
@require_GET
def api_dashboard_filters(request):
    """
    Endpoint para obtener los filtros disponibles.
    Útil para poblar los selectores del frontend.
    """
    filters = DashboardFiltersService.get_available_filters()
    return JsonResponse(filters)


@login_required
@require_GET
def api_dashboard_filtered_stats(request):
    """
    Endpoint para obtener estadísticas filtradas.
    Soporta todos los filtros: fecha, compañía, corredor, tipo, estado.
    """
    filters = DashboardFiltersService.parse_filters_from_request(request)
    stats = DashboardFiltersService.get_filtered_stats(filters)
    return JsonResponse(stats)


@login_required
@require_GET
def api_dashboard_filtered_charts(request):
    """
    Endpoint para obtener datos de gráficos filtrados.
    """
    filters = DashboardFiltersService.parse_filters_from_request(request)
    charts = DashboardFiltersService.get_chart_data(filters)
    return JsonResponse(charts)


@login_required
@require_GET
def api_dashboard_filtered_lists(request):
    """
    Endpoint para obtener listas de registros filtrados.
    Parámetros:
        - limit: Cantidad máxima de registros (default: 5, max: 50)
    """
    filters = DashboardFiltersService.parse_filters_from_request(request)

    try:
        limit = int(request.GET.get("limit", 5))
        limit = min(max(limit, 1), 50)
    except ValueError:
        limit = 5

    lists = DashboardFiltersService.get_lists_data(filters, limit)
    return JsonResponse(lists)


@login_required
@require_GET
def api_dashboard_export(request):
    """
    Endpoint para exportar datos del dashboard en JSON.
    Útil para integraciones y reportes personalizados.
    """
    filters = DashboardFiltersService.parse_filters_from_request(request)
    export_data = DashboardFiltersService.export_filtered_data(filters)

    response = JsonResponse(export_data)
    response["Content-Disposition"] = 'attachment; filename="dashboard_export.json"'
    return response


# Vista personalizada de Login para manejar redirecciones inteligentes
class CustomLoginView(DjangoLoginView):
    template_name = "registration/login.html"

    def get_success_url(self):
        """
        Redirige al usuario según su origen:
        - Si viene del admin (/admin/) → redirige al admin
        - Si viene de la app web → redirige al dashboard
        - Si accede directamente → redirige al dashboard
        """
        # Obtener el parámetro 'next' de la URL
        next_url = self.request.GET.get("next", "")

        # Si hay un 'next' y es del admin, redirigir al admin
        if next_url and ("/admin/" in next_url or next_url.startswith("/admin")):
            return "/admin/"

        # Si hay un 'next' y no es del admin, usar ese
        if next_url and not next_url.startswith("/admin"):
            return next_url

        # Si no hay 'next', verificar el referer
        referer = self.request.META.get("HTTP_REFERER", "")
        if referer and "/admin/" in referer:
            return "/admin/"

        # Por defecto, redirigir al dashboard
        return reverse("dashboard")


# Vista personalizada de Logout que acepta GET
@login_required
def custom_logout(request):
    """
    Vista de logout que acepta tanto GET como POST.
    Hace logout del usuario y redirige al login.
    """
    logout(request)
    return redirect("login")


# =============================================================================
# VISTAS PARA NUEVOS MÓDULOS (Código en inglés, interfaz en español)
# =============================================================================

# --- RENOVACIONES DE PÓLIZAS ---


@login_required
def renewals_list(request):
    """Lista de renovaciones de pólizas con filtros"""
    renewals = PolicyRenewal.objects.select_related(
        "original_policy", "original_policy__compania_aseguradora", "new_policy", "quote", "created_by"
    ).order_by("-due_date")

    # Filtros
    query = request.GET.get("q", "").strip()
    if query:
        renewals = renewals.filter(
            Q(renewal_number__icontains=query) | Q(original_policy__numero_poliza__icontains=query)
        )

    status = request.GET.get("status")
    if status:
        renewals = renewals.filter(status=status)

    decision = request.GET.get("decision")
    if decision:
        renewals = renewals.filter(decision=decision)

    # Paginación
    paginator = Paginator(renewals, 20)
    page = request.GET.get("page", 1)
    renewals_page = paginator.get_page(page)

    # Estadísticas
    stats = {
        "total": PolicyRenewal.objects.count(),
        "pending": PolicyRenewal.objects.filter(status="pending").count(),
        "in_progress": PolicyRenewal.objects.filter(status="in_progress").count(),
        "overdue": PolicyRenewal.objects.filter(
            status__in=["pending", "in_progress"], due_date__lt=timezone.now().date()
        ).count(),
    }

    context = {
        "renewals": renewals_page,
        "stats": stats,
        "status_choices": PolicyRenewal.STATUS_CHOICES,
        "decision_choices": PolicyRenewal.DECISION_CHOICES,
        "query": query,
        "selected_status": status,
        "selected_decision": decision,
    }

    return render(request, "app/renewals/list.html", context)


# --- COTIZACIONES ---


@login_required
def quotes_list(request):
    """Lista de cotizaciones con filtros"""
    quotes = (
        Quote.objects.select_related("policy_type", "resulting_policy", "requested_by")
        .prefetch_related("options")
        .order_by("-request_date")
    )

    # Filtros
    query = request.GET.get("q", "").strip()
    if query:
        quotes = quotes.filter(Q(quote_number__icontains=query) | Q(title__icontains=query))

    status = request.GET.get("status")
    if status:
        quotes = quotes.filter(status=status)

    priority = request.GET.get("priority")
    if priority:
        quotes = quotes.filter(priority=priority)

    # Paginación
    paginator = Paginator(quotes, 20)
    page = request.GET.get("page", 1)
    quotes_page = paginator.get_page(page)

    # Estadísticas
    stats = {
        "total": Quote.objects.count(),
        "draft": Quote.objects.filter(status="draft").count(),
        "sent": Quote.objects.filter(status="sent").count(),
        "accepted": Quote.objects.filter(status="accepted").count(),
        "converted": Quote.objects.filter(status="converted").count(),
    }

    context = {
        "quotes": quotes_page,
        "stats": stats,
        "status_choices": Quote.STATUS_CHOICES,
        "priority_choices": Quote.PRIORITY_CHOICES,
        "query": query,
        "selected_status": status,
        "selected_priority": priority,
    }

    return render(request, "app/quotes/list.html", context)


# --- BIENES ASEGURADOS ---


@login_required
def assets_list(request):
    """Lista de bienes asegurados con filtros (usa modelo unificado BienAsegurado)"""
    assets = BienAsegurado.objects.select_related(
        "poliza", "poliza__compania_aseguradora", "responsable_custodio", "subgrupo_ramo", "creado_por"
    ).order_by("nombre")

    # Filtros
    query = request.GET.get("q", "").strip()
    if query:
        assets = assets.filter(
            Q(codigo_bien__icontains=query)
            | Q(nombre__icontains=query)
            | Q(serie__icontains=query)
            | Q(marca__icontains=query)
        )

    status = request.GET.get("status")
    if status:
        assets = assets.filter(estado=status)

    category = request.GET.get("category")
    if category:
        assets = assets.filter(categoria=category)

    covered = request.GET.get("covered")
    if covered == "yes":
        assets = assets.filter(poliza__isnull=False)
    elif covered == "no":
        assets = assets.filter(poliza__isnull=True)

    # Paginación
    paginator = Paginator(assets, 20)
    page = request.GET.get("page", 1)
    assets_page = paginator.get_page(page)

    # Estadísticas
    from django.db.models import Sum

    stats = {
        "total": BienAsegurado.objects.count(),
        "active": BienAsegurado.objects.filter(estado="activo").count(),
        "covered": BienAsegurado.objects.filter(poliza__isnull=False).count(),
        "not_covered": BienAsegurado.objects.filter(poliza__isnull=True).count(),
        "total_value": BienAsegurado.objects.aggregate(total=Sum("valor_actual"))["total"] or 0,
    }

    # Categorías únicas para el filtro
    categories = BienAsegurado.objects.values_list("categoria", flat=True).distinct().order_by("categoria")

    context = {
        "assets": assets_page,
        "stats": stats,
        "status_choices": BienAsegurado.ESTADO_CHOICES,
        "categories": categories,
        "query": query,
        "selected_status": status,
        "selected_category": category,
        "selected_covered": covered,
    }

    return render(request, "app/assets/list.html", context)


# --- CALENDARIO ---


@login_required
def calendar_view(request):
    """Vista del calendario interactivo"""
    import json
    from datetime import timedelta

    from django.utils import timezone

    today = timezone.now().date()

    # Obtener eventos del mes actual y próximo
    start_range = today.replace(day=1) - timedelta(days=7)
    end_range = (today.replace(day=28) + timedelta(days=35)).replace(day=1) + timedelta(days=6)

    events = CalendarEvent.objects.filter(start_date__gte=start_range, start_date__lte=end_range).select_related(
        "policy", "invoice", "renewal", "claim", "quote"
    )

    # Convertir eventos a formato para FullCalendar
    events_json = []
    for event in events:
        event_data = {
            "id": event.id,
            "title": event.title,
            "start": event.start_date.isoformat(),
            "end": event.end_date.isoformat() if event.end_date else event.start_date.isoformat(),
            "color": event.color,
            "allDay": event.all_day,
            "extendedProps": {
                "type": event.event_type,
                "priority": event.priority,
                "description": event.description,
                "isCompleted": event.is_completed,
            },
        }
        if not event.all_day and event.start_time:
            event_data["start"] = f"{event.start_date.isoformat()}T{event.start_time.isoformat()}"
            if event.end_time:
                end_date = event.end_date or event.start_date
                event_data["end"] = f"{end_date.isoformat()}T{event.end_time.isoformat()}"
        events_json.append(event_data)

    # Próximos eventos (lista)
    upcoming_events = CalendarEvent.objects.filter(start_date__gte=today, is_completed=False).order_by(
        "start_date", "start_time"
    )[:10]

    # Eventos vencidos sin completar
    overdue_events = CalendarEvent.objects.filter(start_date__lt=today, is_completed=False).order_by("-start_date")[:5]

    context = {
        "events_json": json.dumps(events_json),
        "upcoming_events": upcoming_events,
        "overdue_events": overdue_events,
        "event_type_choices": CalendarEvent.EVENT_TYPE_CHOICES,
        "priority_choices": CalendarEvent.PRIORITY_CHOICES,
        "today": today,
    }

    return render(request, "app/calendar/view.html", context)


@login_required
@require_GET
def api_calendar_events(request):
    """API para obtener eventos del calendario (para FullCalendar)"""
    import json
    from datetime import datetime

    start = request.GET.get("start")
    end = request.GET.get("end")

    events = CalendarEvent.objects.all()

    if start:
        start_date = datetime.fromisoformat(start.replace("Z", "+00:00")).date()
        events = events.filter(start_date__gte=start_date)

    if end:
        end_date = datetime.fromisoformat(end.replace("Z", "+00:00")).date()
        events = events.filter(start_date__lte=end_date)

    events_data = []
    for event in events:
        event_data = {
            "id": event.id,
            "title": event.title,
            "start": event.start_date.isoformat(),
            "end": (event.end_date or event.start_date).isoformat(),
            "color": event.color,
            "allDay": event.all_day,
            "url": f"/admin/app/calendarevent/{event.id}/change/",
            "extendedProps": {
                "type": event.get_event_type_display(),
                "priority": event.get_priority_display(),
                "isCompleted": event.is_completed,
            },
        }
        events_data.append(event_data)

    return JsonResponse(events_data, safe=False)


@login_required
def generate_calendar_events(request):
    """Genera eventos automáticos del calendario basados en datos existentes"""
    from django.contrib import messages
    from django.utils import timezone

    today = timezone.now().date()
    created_count = 0

    # Generar eventos para pólizas por vencer
    expiring_policies = Poliza.objects.filter(
        estado__in=["vigente", "por_vencer"], fecha_fin__gte=today, fecha_fin__lte=today + timezone.timedelta(days=60)
    )

    for policy in expiring_policies:
        event, created = CalendarEvent.objects.get_or_create(
            policy=policy,
            event_type="policy_expiry",
            defaults={
                "title": f"Vencimiento: {policy.numero_poliza}",
                "description": f"Vence la póliza {policy.numero_poliza} de {policy.compania_aseguradora}",
                "start_date": policy.fecha_fin,
                "priority": "high" if (policy.fecha_fin - today).days <= 15 else "medium",
                "is_auto_generated": True,
            },
        )
        if created:
            created_count += 1

    # Generar eventos para facturas por vencer
    pending_invoices = Factura.objects.filter(
        estado__in=["pendiente", "parcial"],
        fecha_vencimiento__gte=today,
        fecha_vencimiento__lte=today + timezone.timedelta(days=30),
    )

    for invoice in pending_invoices:
        event, created = CalendarEvent.objects.get_or_create(
            invoice=invoice,
            event_type="invoice_due",
            defaults={
                "title": f"Vence factura: {invoice.numero_factura}",
                "description": f"Vence la factura {invoice.numero_factura} - ${invoice.saldo_pendiente:,.2f}",
                "start_date": invoice.fecha_vencimiento,
                "priority": "high" if (invoice.fecha_vencimiento - today).days <= 7 else "medium",
                "is_auto_generated": True,
            },
        )
        if created:
            created_count += 1

    # Generar eventos para renovaciones pendientes
    pending_renewals = PolicyRenewal.objects.filter(
        status__in=["pending", "in_progress"], due_date__gte=today, due_date__lte=today + timezone.timedelta(days=45)
    )

    for renewal in pending_renewals:
        event, created = CalendarEvent.objects.get_or_create(
            renewal=renewal,
            event_type="renewal_due",
            defaults={
                "title": f"Renovación: {renewal.original_policy.numero_poliza}",
                "description": f"Fecha límite para renovar la póliza {renewal.original_policy.numero_poliza}",
                "start_date": renewal.due_date,
                "priority": "critical" if (renewal.due_date - today).days <= 7 else "high",
                "is_auto_generated": True,
            },
        )
        if created:
            created_count += 1

    messages.success(request, f"Se generaron {created_count} eventos automáticos.")
    return redirect("calendar_view")


# --- APROBACIONES DE PAGO ---


@login_required
def payment_approvals_list(request):
    """Lista de aprobaciones de pago pendientes"""
    approvals = PaymentApproval.objects.select_related(
        "payment", "payment__factura", "payment__factura__poliza", "approver"
    ).order_by("-requested_at")

    # Filtros
    status = request.GET.get("status")
    if status:
        approvals = approvals.filter(status=status)
    else:
        # Por defecto mostrar solo pendientes
        approvals = approvals.filter(status="pending")

    level = request.GET.get("level")
    if level:
        approvals = approvals.filter(required_level=level)

    # Paginación
    paginator = Paginator(approvals, 20)
    page = request.GET.get("page", 1)
    approvals_page = paginator.get_page(page)

    # Estadísticas
    stats = {
        "pending": PaymentApproval.objects.filter(status="pending").count(),
        "approved_today": PaymentApproval.objects.filter(
            status="approved", decided_at__date=timezone.now().date()
        ).count(),
        "rejected_today": PaymentApproval.objects.filter(
            status="rejected", decided_at__date=timezone.now().date()
        ).count(),
    }

    context = {
        "approvals": approvals_page,
        "stats": stats,
        "status_choices": PaymentApproval.STATUS_CHOICES,
        "level_choices": PaymentApproval.APPROVAL_LEVEL_CHOICES,
        "selected_status": status or "pending",
        "selected_level": level,
    }

    return render(request, "app/approvals/list.html", context)


@login_required
@require_POST
def approve_payment(request, pk):
    """Aprobar un pago"""
    from django.contrib import messages

    approval = get_object_or_404(PaymentApproval, pk=pk)

    if approval.status != "pending":
        messages.error(request, "Esta aprobación ya fue procesada.")
        return redirect("payment_approvals_list")

    notes = request.POST.get("notes", "")
    approval.approve(request.user, notes)

    messages.success(request, "Pago aprobado correctamente.")
    return redirect("payment_approvals_list")


@login_required
@require_POST
def reject_payment(request, pk):
    """Rechazar un pago"""
    from django.contrib import messages

    approval = get_object_or_404(PaymentApproval, pk=pk)

    if approval.status != "pending":
        messages.error(request, "Esta aprobación ya fue procesada.")
        return redirect("payment_approvals_list")

    notes = request.POST.get("notes", "")
    if not notes:
        messages.error(request, "Debe proporcionar un motivo de rechazo.")
        return redirect("payment_approvals_list")

    approval.reject(request.user, notes)

    messages.warning(request, "Pago rechazado.")
    return redirect("payment_approvals_list")


# =============================================================================
# DASHBOARD ESPECIALIZADO DE ANALYTICS
# =============================================================================


@login_required
def analytics_dashboard(request):
    """
    Dashboard especializado con analytics avanzados:
    - Ratio de siniestralidad por tipo de póliza
    - Tendencia indemnizaciones vs primas
    - Heatmap de ubicaciones
    """
    import json

    # Obtener datos del servicio
    summary = AdvancedAnalyticsService.get_dashboard_summary()
    loss_ratio_data = AdvancedAnalyticsService.get_loss_ratio_by_policy_type()
    trend_data = AdvancedAnalyticsService.get_claims_vs_premiums_trend(months=12)
    location_data = AdvancedAnalyticsService.get_claims_by_location()
    claims_distribution = AdvancedAnalyticsService.get_claims_by_type_distribution()
    insurer_performance = AdvancedAnalyticsService.get_insurer_performance()

    context = {
        "summary": summary,
        "loss_ratio_data": loss_ratio_data,
        "trend_data_json": json.dumps(trend_data),
        "location_data": location_data,
        "location_data_json": json.dumps(location_data),
        "claims_distribution_json": json.dumps(claims_distribution),
        "insurer_performance": insurer_performance,
    }

    return render(request, "app/analytics/dashboard.html", context)


@login_required
@require_GET
def api_analytics_loss_ratio(request):
    """API para obtener ratio de siniestralidad"""
    data = AdvancedAnalyticsService.get_loss_ratio_by_policy_type()
    return JsonResponse({"data": data})


@login_required
@require_GET
def api_analytics_trend(request):
    """API para obtener tendencia de primas vs indemnizaciones"""
    months = int(request.GET.get("months", 12))
    data = AdvancedAnalyticsService.get_claims_vs_premiums_trend(months=months)
    return JsonResponse(data)


@login_required
@require_GET
def api_analytics_locations(request):
    """API para obtener datos de ubicaciones"""
    data = AdvancedAnalyticsService.get_claims_by_location()
    return JsonResponse({"data": data})


@login_required
@require_GET
def api_analytics_insurers(request):
    """API para obtener rendimiento de aseguradoras"""
    data = AdvancedAnalyticsService.get_insurer_performance()
    return JsonResponse({"data": data})


# =============================================================================
# VISTAS BASADAS EN CLASES - RAMOS
# =============================================================================


class RamoListView(LoginRequiredMixin, ListView):
    """Lista de ramos"""

    model = Ramo
    template_name = "app/ramos/lista.html"
    context_object_name = "ramos"
    paginate_by = 20

    def get_queryset(self):
        queryset = Ramo.objects.all().order_by("nombre")
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(Q(codigo__icontains=query) | Q(nombre__icontains=query))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")
        return context


class RamoCreateView(LoginRequiredMixin, CreateView):
    """Crear ramo"""

    model = Ramo
    form_class = RamoForm
    template_name = "app/ramos/crear.html"
    success_url = reverse_lazy("ramos_lista")

    def form_valid(self, form):
        messages.success(self.request, "Ramo creado exitosamente.")
        return super().form_valid(form)


class RamoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar ramo"""

    model = Ramo
    form_class = RamoForm
    template_name = "app/ramos/editar.html"
    success_url = reverse_lazy("ramos_lista")

    def form_valid(self, form):
        messages.success(self.request, "Ramo actualizado exitosamente.")
        return super().form_valid(form)


# =============================================================================
# VISTAS RÁPIDAS - CREAR ENTIDADES BASE (popup/modal)
# =============================================================================


class CompaniaAseguradoraCreateView(LoginRequiredMixin, CreateView):
    """Crear compañía aseguradora"""

    model = CompaniaAseguradora
    form_class = CompaniaAseguradoraForm
    template_name = "app/components/crear_rapido.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nueva Compañía Aseguradora"
        context["icono"] = "fas fa-building"
        context["volver_url"] = self.request.GET.get("next", reverse_lazy("poliza_crear"))
        return context

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("poliza_crear")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Compañía aseguradora "{self.object.nombre}" creada exitosamente.')
        return response


class CorredorSegurosCreateView(LoginRequiredMixin, CreateView):
    """Crear corredor de seguros"""

    model = CorredorSeguros
    form_class = CorredorSegurosForm
    template_name = "app/components/crear_rapido.html"

    def get_initial(self):
        initial = super().get_initial()
        # Preseleccionar la compañía si viene en los parámetros
        compania_id = self.request.GET.get("compania_id")
        if compania_id:
            initial["compania_aseguradora"] = compania_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Corredor de Seguros"
        context["icono"] = "fas fa-user-tie"
        context["volver_url"] = self.request.GET.get("next", reverse_lazy("poliza_crear"))
        # Mostrar nombre de la compañía si viene preseleccionada
        compania_id = self.request.GET.get("compania_id")
        if compania_id:
            try:
                compania = CompaniaAseguradora.objects.get(pk=compania_id)
                context["subtitulo"] = f"Para: {compania.nombre}"
            except CompaniaAseguradora.DoesNotExist:
                pass
        return context

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("poliza_crear")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Corredor de seguros "{self.object.nombre}" creado exitosamente.')
        return response


class TipoSiniestroCreateView(LoginRequiredMixin, CreateView):
    """Crear tipo de siniestro"""

    model = TipoSiniestro
    form_class = TipoSiniestroForm
    template_name = "app/components/crear_rapido.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Tipo de Siniestro"
        context["icono"] = "fas fa-exclamation-triangle"
        context["volver_url"] = self.request.GET.get("next", reverse_lazy("siniestro_crear"))
        return context

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("siniestro_crear")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Tipo de siniestro "{self.object.nombre}" creado exitosamente.')
        return response


class ResponsableCustodioCreateView(LoginRequiredMixin, CreateView):
    """Crear responsable/custodio"""

    model = ResponsableCustodio
    form_class = ResponsableCustodioForm
    template_name = "app/components/crear_rapido.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Responsable/Custodio"
        context["icono"] = "fas fa-user-shield"
        context["volver_url"] = self.request.GET.get("next", reverse_lazy("siniestro_crear"))
        return context

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("siniestro_crear")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Responsable/Custodio "{self.object.nombre}" creado exitosamente.')
        return response


# =============================================================================
# VISTAS BASADAS EN CLASES - PÓLIZAS (CRUD Completo)
# =============================================================================


class PolizaCreateView(LoginRequiredMixin, CreateView):
    """Crear póliza con formset de ramos"""

    model = Poliza
    form_class = PolizaForm
    template_name = "app/polizas/crear.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["detalles_formset"] = DetallePolizaRamoFormSet(self.request.POST)
        else:
            context["detalles_formset"] = DetallePolizaRamoFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        detalles_formset = context["detalles_formset"]

        with transaction.atomic():
            form.instance.creado_por = self.request.user
            # Asignar automáticamente el tipo de póliza "Ramos Generales"
            tipo_ramos_generales, _ = TipoPoliza.objects.get_or_create(
                nombre="Ramos Generales", defaults={"descripcion": "Tipo de póliza por defecto", "activo": True}
            )
            form.instance.tipo_poliza = tipo_ramos_generales
            self.object = form.save()

            if detalles_formset.is_valid():
                detalles_formset.instance = self.object
                detalles_formset.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, f"Póliza {self.object.numero_poliza} creada exitosamente.")
        return redirect("poliza_detalle", pk=self.object.pk)


class PolizaUpdateView(LoginRequiredMixin, UpdateView):
    """Editar póliza con formset de ramos"""

    model = Poliza
    form_class = PolizaForm
    template_name = "app/polizas/editar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["detalles_formset"] = DetallePolizaRamoFormSet(self.request.POST, instance=self.object)
        else:
            context["detalles_formset"] = DetallePolizaRamoFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        detalles_formset = context["detalles_formset"]

        with transaction.atomic():
            self.object = form.save()

            if detalles_formset.is_valid():
                detalles_formset.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, f"Póliza {self.object.numero_poliza} actualizada exitosamente.")
        return redirect("poliza_detalle", pk=self.object.pk)


class PolizaDetailView(LoginRequiredMixin, DetailView):
    """Detalle de póliza con tabla de ramos"""

    model = Poliza
    template_name = "app/polizas/detalle.html"
    context_object_name = "poliza"

    def get_queryset(self):
        """Optimización: Pre-calcular total_prima_ramos con annotate()"""
        from django.db.models import Sum

        return super().get_queryset().annotate(total_prima_calculado=Sum("detalles_ramo__total_prima"))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Usar select_related para evitar N+1 en ramos
        context["detalles_ramo"] = self.object.detalles_ramo.select_related("subgrupo_ramo")
        context["siniestros"] = self.object.siniestros.select_related("tipo_siniestro").order_by("-fecha_siniestro")[:5]
        context["facturas"] = self.object.facturas.order_by("-fecha_emision")[:5]
        # Usar el valor pre-calculado si está disponible
        context["total_prima_ramos"] = (
            getattr(self.object, "total_prima_calculado", None) or self.object.total_prima_ramos
        )
        return context


# =============================================================================
# VISTAS BASADAS EN CLASES - SINIESTROS (CRUD Completo)
# =============================================================================


class SiniestroCreateView(LoginRequiredMixin, CreateView):
    """Crear siniestro con adjuntos y checklist"""

    model = Siniestro
    form_class = SiniestroForm
    template_name = "app/siniestros/crear.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["adjuntos_formset"] = AdjuntoSiniestroFormSet(self.request.POST, self.request.FILES)
        else:
            context["adjuntos_formset"] = AdjuntoSiniestroFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        adjuntos_formset = context["adjuntos_formset"]

        with transaction.atomic():
            form.instance.creado_por = self.request.user
            self.object = form.save()

            if adjuntos_formset.is_valid():
                adjuntos_formset.instance = self.object
                for adjunto_form in adjuntos_formset:
                    if adjunto_form.cleaned_data and not adjunto_form.cleaned_data.get("DELETE"):
                        adjunto = adjunto_form.save(commit=False)
                        adjunto.siniestro = self.object
                        adjunto.subido_por = self.request.user
                        adjunto.save()

            # Crear checklist inicial basado en tipo de siniestro
            self._crear_checklist_inicial()

        messages.success(self.request, f"Siniestro {self.object.numero_siniestro} registrado exitosamente.")
        return redirect("siniestro_detalle", pk=self.object.pk)

    def _crear_checklist_inicial(self):
        """Crea el checklist inicial basado en el tipo de siniestro"""
        if self.object.tipo_siniestro:
            items_config = ChecklistSiniestroConfig.objects.filter(
                tipo_siniestro=self.object.tipo_siniestro, activo=True
            )
            for config in items_config:
                ChecklistSiniestro.objects.create(siniestro=self.object, config_item=config, completado=False)


class SiniestroUpdateView(LoginRequiredMixin, UpdateView):
    """Editar siniestro"""

    model = Siniestro
    form_class = SiniestroForm
    template_name = "app/siniestros/editar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["adjuntos_formset"] = AdjuntoSiniestroFormSet(
                self.request.POST, self.request.FILES, instance=self.object
            )
            context["gestion_form"] = SiniestroGestionForm(self.request.POST, instance=self.object)
        else:
            context["adjuntos_formset"] = AdjuntoSiniestroFormSet(instance=self.object)
            context["gestion_form"] = SiniestroGestionForm(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        adjuntos_formset = context["adjuntos_formset"]
        gestion_form = context["gestion_form"]

        with transaction.atomic():
            self.object = form.save()

            # Guardar campos de gestión
            if gestion_form.is_valid():
                for field in gestion_form.cleaned_data:
                    setattr(self.object, field, gestion_form.cleaned_data[field])
                self.object.save()

            if adjuntos_formset.is_valid():
                for adjunto_form in adjuntos_formset:
                    if adjunto_form.cleaned_data:
                        if adjunto_form.cleaned_data.get("DELETE") and adjunto_form.instance.pk:
                            adjunto_form.instance.delete()
                        elif not adjunto_form.cleaned_data.get("DELETE"):
                            adjunto = adjunto_form.save(commit=False)
                            adjunto.siniestro = self.object
                            if not adjunto.pk:
                                adjunto.subido_por = self.request.user
                            adjunto.save()

        messages.success(self.request, f"Siniestro {self.object.numero_siniestro} actualizado exitosamente.")
        return redirect("siniestro_detalle", pk=self.object.pk)


class SiniestroDetailView(LoginRequiredMixin, DetailView):
    """Detalle de siniestro con timeline y checklist interactivo"""

    model = Siniestro
    template_name = "app/siniestros/detalle.html"
    context_object_name = "siniestro"

    # Mapeo de estados a índice numérico para la barra de progreso
    # Agrupados: 1-Registro, 2-Gestión, 3-Enviado, 4-Recibo, 5-Liquidación, 6-Cerrado
    ESTADO_INDEX = {
        "registrado": 1,
        "notificado_broker": 2,
        "documentacion_lista": 2,
        "enviado_aseguradora": 3,
        "recibo_recibido": 4,
        "en_disputa": 4,
        "recibo_firmado": 4,
        "pendiente_liquidacion": 5,
        "vencido": 5,
        "liquidado": 5,
        "cerrado": 6,
        "rechazado": 0,  # Estado especial
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Adjuntos
        adjuntos = self.object.adjuntos.all()
        context["adjuntos"] = adjuntos

        # Checklist con adjuntos vinculados
        checklist_items = list(self.object.checklist_items.select_related("config_item", "completado_por"))

        # Vincular adjuntos a items del checklist (usar nombre diferente al related_name)
        for item in checklist_items:
            item.adjuntos_del_item = list(adjuntos.filter(checklist_item=item))

        context["checklist"] = checklist_items

        # Estadísticas del checklist
        total_items = len(checklist_items)
        completados = sum(1 for item in checklist_items if item.completado)
        obligatorios = [item for item in checklist_items if item.config_item.es_obligatorio]
        obligatorios_completados = sum(1 for item in obligatorios if item.completado)
        obligatorios_pendientes = len(obligatorios) - obligatorios_completados

        context["checklist_total"] = total_items
        context["checklist_completados"] = completados
        context["checklist_porcentaje"] = (completados * 100 // total_items) if total_items > 0 else 0
        context["checklist_obligatorios_pendientes"] = obligatorios_pendientes
        context["checklist_completo"] = obligatorios_pendientes == 0 and len(obligatorios) > 0

        # Timeline
        context["timeline"] = self._generar_timeline()

        # Alertas
        context["alertas"] = {
            "respuesta": self.object.alerta_respuesta_aseguradora,
            "responsable": self.object.alerta_notificar_responsable,
            "deposito": self.object.alerta_deposito_pendiente,
        }

        # Índice del estado actual para la barra de progreso
        context["estado_index"] = self.ESTADO_INDEX.get(self.object.estado, 1)

        return context

    def _generar_timeline(self):
        """Genera un timeline de eventos del siniestro"""
        from datetime import date, datetime

        from django.utils import timezone

        def to_datetime(d):
            """Convierte date a datetime naive para comparación uniforme"""
            if d is None:
                return None
            if isinstance(d, datetime):
                # Si es aware, convertir a naive en UTC
                if timezone.is_aware(d):
                    return timezone.make_naive(d)
                return d
            if isinstance(d, date):
                return datetime.combine(d, datetime.min.time())
            return d

        timeline = []
        s = self.object

        if s.fecha_registro:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_registro),
                    "titulo": "Siniestro Registrado",
                    "tipo": "registro",
                    "icono": "file-text",
                }
            )

        if s.fecha_notificacion_broker:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_notificacion_broker),
                    "titulo": "Notificado al Broker",
                    "tipo": "notificacion",
                    "icono": "send",
                }
            )

        if s.fecha_envio_aseguradora:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_envio_aseguradora),
                    "titulo": "Enviado a Aseguradora",
                    "tipo": "envio",
                    "icono": "upload",
                }
            )

        if s.fecha_respuesta_aseguradora:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_respuesta_aseguradora),
                    "titulo": "Respuesta de Aseguradora",
                    "tipo": "respuesta",
                    "icono": "check-circle",
                }
            )

        if s.fecha_liquidacion:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_liquidacion),
                    "titulo": "Siniestro Liquidado",
                    "tipo": "liquidacion",
                    "icono": "dollar-sign",
                }
            )

        if s.fecha_firma_indemnizacion:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_firma_indemnizacion),
                    "titulo": "Recibo de Indemnización Firmado",
                    "tipo": "firma",
                    "icono": "edit",
                }
            )

        if s.fecha_pago:
            timeline.append(
                {
                    "fecha": to_datetime(s.fecha_pago),
                    "titulo": "Pago Realizado",
                    "tipo": "pago",
                    "icono": "credit-card",
                }
            )

        return sorted(timeline, key=lambda x: x["fecha"]) if timeline else []


# =============================================================================
# VISTAS BASADAS EN CLASES - GRUPOS DE BIENES
# =============================================================================


class GrupoBienesListView(LoginRequiredMixin, ListView):
    """Lista de grupos de bienes"""

    model = GrupoBienes
    template_name = "app/grupos_bienes/lista.html"
    context_object_name = "grupos"
    paginate_by = 20

    def get_queryset(self):
        return GrupoBienes.objects.select_related("ramo", "responsable", "poliza").order_by("nombre")


class GrupoBienesCreateView(LoginRequiredMixin, CreateView):
    """Crear grupo de bienes"""

    model = GrupoBienes
    form_class = GrupoBienesForm
    template_name = "app/grupos_bienes/crear.html"
    success_url = reverse_lazy("grupos_bienes_lista")

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, "Grupo de bienes creado exitosamente.")
        return super().form_valid(form)


class GrupoBienesDetailView(LoginRequiredMixin, DetailView):
    """Detalle de grupo de bienes"""

    model = GrupoBienes
    template_name = "app/grupos_bienes/detalle.html"
    context_object_name = "grupo"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["bienes"] = self.object.bienes.all()
        return context


# =============================================================================
# VISTAS BASADAS EN CLASES - BIENES ASEGURADOS (Modelo Unificado)
# =============================================================================


class BienAseguradoCreateView(LoginRequiredMixin, CreateView):
    """Crear bien asegurado (modelo unificado BienAsegurado)"""

    model = BienAsegurado
    form_class = BienAseguradoForm
    template_name = "app/bienes/crear.html"
    success_url = reverse_lazy("assets_list")

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, "Bien asegurado creado exitosamente.")
        return super().form_valid(form)


class BienAseguradoUpdateView(LoginRequiredMixin, UpdateView):
    """Editar bien asegurado (modelo unificado BienAsegurado)"""

    model = BienAsegurado
    form_class = BienAseguradoForm
    template_name = "app/bienes/editar.html"
    success_url = reverse_lazy("assets_list")

    def form_valid(self, form):
        messages.success(self.request, "Bien asegurado actualizado exitosamente.")
        return super().form_valid(form)


class BienAseguradoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de bien asegurado (modelo unificado BienAsegurado)"""

    model = BienAsegurado
    template_name = "app/bienes/detalle.html"
    context_object_name = "bien"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agregar siniestros relacionados
        context["siniestros"] = self.object.siniestros.select_related("tipo_siniestro").order_by("-fecha_siniestro")[:5]
        return context


# =============================================================================
# VISTAS DE REPORTES AVANZADOS
# =============================================================================


@login_required
def reporte_siniestralidad(request):
    """Vista del reporte de siniestralidad"""
    from datetime import timedelta

    fecha_hasta = timezone.now().date()
    fecha_desde = fecha_hasta - timedelta(days=365)

    if request.GET.get("fecha_desde"):
        from datetime import datetime

        fecha_desde = datetime.strptime(request.GET["fecha_desde"], "%Y-%m-%d").date()
    if request.GET.get("fecha_hasta"):
        from datetime import datetime

        fecha_hasta = datetime.strptime(request.GET["fecha_hasta"], "%Y-%m-%d").date()

    compania_id = request.GET.get("compania")

    datos = ReportesAvanzadosService.calcular_siniestralidad(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        compania_id=compania_id,
    )

    context = {
        "datos": datos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "companias": CompaniaAseguradora.objects.filter(activo=True),
        "compania_seleccionada": compania_id,
    }

    return render(request, "app/reportes/siniestralidad.html", context)


@login_required
def reporte_gasto_ramos(request):
    """Vista del reporte de gastos por ramos"""
    from datetime import timedelta

    fecha_hasta = timezone.now().date()
    fecha_desde = fecha_hasta - timedelta(days=365)

    if request.GET.get("fecha_desde"):
        from datetime import datetime

        fecha_desde = datetime.strptime(request.GET["fecha_desde"], "%Y-%m-%d").date()
    if request.GET.get("fecha_hasta"):
        from datetime import datetime

        fecha_hasta = datetime.strptime(request.GET["fecha_hasta"], "%Y-%m-%d").date()

    datos = ReportesAvanzadosService.reporte_gasto_por_ramos(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    context = {
        "datos": datos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }

    return render(request, "app/reportes/gasto_ramos.html", context)


@login_required
def reporte_dias_gestion(request):
    """Vista del reporte de días de gestión"""
    from datetime import timedelta

    fecha_hasta = timezone.now().date()
    fecha_desde = fecha_hasta - timedelta(days=365)

    if request.GET.get("fecha_desde"):
        from datetime import datetime

        fecha_desde = datetime.strptime(request.GET["fecha_desde"], "%Y-%m-%d").date()
    if request.GET.get("fecha_hasta"):
        from datetime import datetime

        fecha_hasta = datetime.strptime(request.GET["fecha_hasta"], "%Y-%m-%d").date()

    datos = ReportesAvanzadosService.reporte_dias_gestion_siniestros(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    context = {
        "datos": datos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }

    return render(request, "app/reportes/dias_gestion.html", context)


@login_required
def reporte_siniestros_dependencia(request):
    """Vista del reporte de siniestros por dependencia"""
    from datetime import timedelta

    fecha_hasta = timezone.now().date()
    fecha_desde = fecha_hasta - timedelta(days=365)

    if request.GET.get("fecha_desde"):
        from datetime import datetime

        fecha_desde = datetime.strptime(request.GET["fecha_desde"], "%Y-%m-%d").date()
    if request.GET.get("fecha_hasta"):
        from datetime import datetime

        fecha_hasta = datetime.strptime(request.GET["fecha_hasta"], "%Y-%m-%d").date()

    datos = ReportesAvanzadosService.reporte_siniestros_por_dependencia(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    context = {
        "datos": datos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }

    return render(request, "app/reportes/siniestros_dependencia.html", context)


# =============================================================================
# VISTAS DE ACCIONES DE SINIESTROS
# =============================================================================


@login_required
@require_GET
def siniestro_email_preview(request, pk, tipo):
    """
    API para obtener la vista previa del email antes de enviarlo.
    Retorna JSON con los datos del email para mostrar en el modal.

    Tipos soportados:
    - broker: Notificación al broker
    - aseguradora: Envío a aseguradora
    - responsable: Notificación al responsable del bien
    """
    from django.http import JsonResponse

    siniestro = get_object_or_404(Siniestro, pk=pk)

    try:
        if tipo == "broker":
            data = _generar_preview_broker(siniestro)
        elif tipo == "aseguradora":
            data = _generar_preview_aseguradora(siniestro)
        elif tipo == "responsable":
            data = _generar_preview_responsable(siniestro)
        else:
            return JsonResponse({"success": False, "error": "Tipo de email no válido"})

        data["success"] = True
        data["submit_url"] = reverse("siniestro_email_enviar", args=[pk, tipo])
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def _generar_preview_broker(siniestro):
    """Genera la vista previa del email para el broker"""
    from app.models import ConfiguracionSistema

    # Obtener email del broker
    email_broker = siniestro.email_broker
    if not email_broker and siniestro.poliza and siniestro.poliza.corredor_seguros:
        email_broker = siniestro.poliza.corredor_seguros.email

    if not email_broker:
        raise ValueError("No se encontró email del broker")

    _ = siniestro.get_poliza()
    bien = siniestro.bien_asegurado

    # Datos del bien
    responsable = bien.responsable_custodio.nombre if bien and bien.responsable_custodio else "N/A"
    periferico = bien.nombre if bien else "N/A"
    marca = bien.marca if bien else "N/A"
    modelo = bien.modelo if bien else "N/A"
    serie = bien.serie if bien else "N/A"
    activo = bien.codigo_bien if bien else "N/A"

    # Fecha de reporte
    fecha_reporte = siniestro.fecha_siniestro.strftime("%d/%m/%Y") if siniestro.fecha_siniestro else "N/A"

    # Problema y causa
    problema = siniestro.descripcion_detallada or "N/A"
    causa = siniestro.causa or "N/A"

    # Firmante dinámico desde configuración
    firmante_nombre = ConfiguracionSistema.get_config("FIRMANTE_CARTA_NOMBRE", "Gestión de Seguros")
    firmante_cargo = ConfiguracionSistema.get_config("FIRMANTE_CARTA_CARGO", "Seguros / Inventarios")
    firmante_email = ConfiguracionSistema.get_config("FIRMANTE_EMAIL", "")
    firmante_telefono = ConfiguracionSistema.get_config("FIRMANTE_TELEFONO", "")
    firmante_departamento = ConfiguracionSistema.get_config("FIRMANTE_DEPARTAMENTO", "")

    # Construir firma
    firma_lines = [firmante_nombre, firmante_cargo]
    if firmante_departamento:
        firma_lines.append(firmante_departamento)
    if firmante_telefono:
        firma_lines.append(f"Telf. {firmante_telefono}")
    if firmante_email:
        firma_lines.append(firmante_email)
    firma = "\n".join(firma_lines)

    contenido = f"""Estimado/a,

Se reporta el siguiente siniestro para su gestión:


REPORTE DE SINIESTRO
====================

Usuario responsable: {responsable}

Fecha de reporte a MST (Siniestro): {fecha_reporte}

Problema: {problema}

Causa Probable: {causa}


Los datos del equipo son:

Periférico: {periferico}
Marca:      {marca}
Modelo:     {modelo}
Serie:      {serie}
Activo:     {activo}


{firma}"""

    return {
        "tipo_descripcion": "Notificación al Broker",
        "destinatario": email_broker,
        "cc": "",
        "asunto": f"Notificación de Siniestro - {siniestro.numero_siniestro}",
        "contenido": contenido,
        "adjuntos": [],
    }


def _generar_preview_aseguradora(siniestro):
    """Genera la vista previa del email para la aseguradora"""
    poliza = siniestro.get_poliza()

    if not poliza or not poliza.compania_aseguradora:
        raise ValueError("El siniestro no tiene póliza o aseguradora asociada")

    email_aseguradora = poliza.compania_aseguradora.email
    if not email_aseguradora:
        raise ValueError("La compañía aseguradora no tiene email registrado")

    # Obtener adjuntos
    adjuntos = siniestro.adjuntos.all()
    adjuntos_lista = [{"nombre": adj.nombre, "tipo": adj.get_tipo_adjunto_display()} for adj in adjuntos]

    contenido = f"""Estimados señores,

Por medio de la presente, nos permitimos enviar la documentación correspondiente al siniestro:

DATOS DEL SINIESTRO
===================
Número de Siniestro: {siniestro.numero_siniestro}
Póliza: {poliza.numero_poliza}
Tipo de Siniestro: {siniestro.tipo_siniestro.get_nombre_display() if siniestro.tipo_siniestro else 'N/A'}
Fecha del Siniestro: {siniestro.fecha_siniestro.strftime('%d/%m/%Y') if siniestro.fecha_siniestro else 'N/A'}
Bien Afectado: {siniestro.get_nombre_bien()}
Monto Estimado: ${siniestro.monto_estimado:,.2f}

CAUSA DEL SINIESTRO
===================
{siniestro.causa}

DESCRIPCIÓN DETALLADA
=====================
{siniestro.descripcion_detallada}

DOCUMENTOS ADJUNTOS
===================
"""

    for adj in adjuntos:
        contenido += f"• {adj.nombre} ({adj.get_tipo_adjunto_display()})\n"

    if not adjuntos:
        contenido += "(Sin documentos adjuntos)\n"

    contenido += """
Quedamos atentos a su respuesta dentro del plazo establecido.

Atentamente,
Gestión de Seguros"""

    return {
        "tipo_descripcion": "Envío a Aseguradora",
        "destinatario": email_aseguradora,
        "cc": siniestro.email_broker or "",
        "asunto": f"Siniestro {siniestro.numero_siniestro} - {poliza.numero_poliza}",
        "contenido": contenido,
        "adjuntos": adjuntos_lista,
    }


def _generar_preview_responsable(siniestro):
    """Genera la vista previa del email para el responsable del bien"""
    responsable = siniestro.responsable_custodio
    if not responsable:
        # Intentar obtener del bien asegurado
        if siniestro.bien_asegurado and siniestro.bien_asegurado.responsable_custodio:
            responsable = siniestro.bien_asegurado.responsable_custodio

    if not responsable or not responsable.email:
        raise ValueError("No se encontró email del responsable del bien")

    contenido = f"""Estimado/a {responsable.nombre},

Le informamos sobre el estado del siniestro registrado para el bien bajo su custodia.

INFORMACIÓN DEL SINIESTRO
=========================
Número de Siniestro: {siniestro.numero_siniestro}
Estado Actual: {siniestro.get_estado_display()}
Fecha del Siniestro: {siniestro.fecha_siniestro.strftime('%d/%m/%Y') if siniestro.fecha_siniestro else 'N/A'}

INFORMACIÓN DEL BIEN
====================
Bien: {siniestro.get_nombre_bien()}
Ubicación: {siniestro.ubicacion}

DESCRIPCIÓN
===========
{siniestro.causa}

ESTADO DE LA GESTIÓN
====================
"""

    if siniestro.fecha_envio_aseguradora:
        contenido += f"• Enviado a aseguradora: {siniestro.fecha_envio_aseguradora.strftime('%d/%m/%Y')}\n"
    if siniestro.fecha_respuesta_aseguradora:
        contenido += f"• Respuesta recibida: {siniestro.fecha_respuesta_aseguradora.strftime('%d/%m/%Y')}\n"
    if siniestro.monto_indemnizado:
        contenido += f"• Monto indemnizado: ${siniestro.monto_indemnizado:,.2f}\n"

    contenido += """
Si requiere información adicional, no dude en contactarnos.

Atentamente,
Gestión de Seguros"""

    return {
        "tipo_descripcion": "Notificación al Responsable",
        "destinatario": responsable.email,
        "cc": "",
        "asunto": f"Estado de Siniestro - {siniestro.numero_siniestro}",
        "contenido": contenido,
        "adjuntos": [],
    }


def _generar_html_email(siniestro, tipo, asunto, contenido_texto, notas_adicionales=None):
    """
    Genera el contenido HTML del email usando la plantilla moderna base_notificacion.html.
    """
    from django.conf import settings
    from django.template.loader import render_to_string
    from django.utils import timezone

    from app.models import ConfiguracionSistema

    _ = siniestro.get_poliza()
    bien = siniestro.bien_asegurado

    # Determinar tipo de notificación para mostrar
    tipos_notificacion = {
        "broker": "Notificación al Broker",
        "aseguradora": "Envío a Aseguradora",
        "responsable": "Notificación al Responsable",
    }

    # Datos del bien (priorizar bien_asegurado sobre campos legacy)
    if bien:
        responsable = bien.responsable_custodio.nombre if bien.responsable_custodio else "N/A"
        bien_nombre = bien.nombre
        marca = bien.marca or "N/A"
        modelo = bien.modelo or "N/A"
        serie = bien.serie or "N/A"
        codigo_activo = bien.codigo_bien or "N/A"
    else:
        responsable = str(siniestro.responsable_custodio) if siniestro.responsable_custodio else "N/A"
        bien_nombre = siniestro.get_nombre_bien()
        marca = siniestro.bien_marca or "N/A"
        modelo = siniestro.bien_modelo or "N/A"
        serie = siniestro.bien_serie or "N/A"
        codigo_activo = siniestro.bien_codigo_activo or "N/A"

    # Fecha de reporte
    fecha_reporte = siniestro.fecha_siniestro.strftime("%d/%m/%Y") if siniestro.fecha_siniestro else "N/A"

    # Firmante dinámico
    firmante_nombre = ConfiguracionSistema.get_config("FIRMANTE_CARTA_NOMBRE", "Gestión de Seguros")
    firmante_cargo = ConfiguracionSistema.get_config("FIRMANTE_CARTA_CARGO", "Seguros / Inventarios")
    firmante_departamento = ConfiguracionSistema.get_config("FIRMANTE_DEPARTAMENTO", "")
    firmante_telefono = ConfiguracionSistema.get_config("FIRMANTE_TELEFONO", "")
    firmante_email = ConfiguracionSistema.get_config("FIRMANTE_EMAIL", "")

    # Construir contexto para la plantilla
    context = {
        # Header info
        "tipo_notificacion": tipos_notificacion.get(tipo, "Notificación de Siniestro"),
        # Saludo y mensaje intro
        "saludo": "Estimado/a,",
        "mensaje_intro": "Se reporta el siguiente siniestro para su gestión:",
        # Datos del reporte
        "responsable": responsable,
        "fecha_reporte": fecha_reporte,
        "problema": siniestro.descripcion_detallada or "N/A",
        "causa": siniestro.causa or "N/A",
        # Datos del equipo
        "bien_nombre": bien_nombre,
        "marca": marca,
        "modelo": modelo,
        "serie": serie,
        "codigo_activo": codigo_activo,
        # Notas adicionales
        "notas_adicionales": notas_adicionales,
        # Firma
        "firmante_nombre": firmante_nombre,
        "firmante_cargo": firmante_cargo,
        "firmante_departamento": firmante_departamento,
        "firmante_telefono": firmante_telefono,
        "firmante_email": firmante_email,
        # Footer
        "timestamp": timezone.now().isoformat(),
    }

    # Renderizar plantilla HTML
    return render_to_string("emails/base_notificacion.html", context)


def _crear_email_con_logo(asunto, contenido_texto, contenido_html, destinatario, cc_list=None):
    """
    Crea un EmailMultiAlternatives con el logo de UTPL embebido.
    Esta función centraliza la lógica de adjuntar imágenes a emails.
    """
    import os
    from email.mime.image import MIMEImage

    from django.conf import settings
    from django.core.mail import EmailMultiAlternatives

    email = EmailMultiAlternatives(
        subject=asunto,
        body=contenido_texto,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[destinatario],
        cc=cc_list if cc_list else None,
    )

    # Agregar versión HTML
    email.attach_alternative(contenido_html, "text/html")

    # Adjuntar logo como imagen embebida con CID
    logo_path = os.path.join(settings.BASE_DIR, "app", "static", "images", "logo-utpl.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_data = f.read()
        logo_image = MIMEImage(logo_data)
        logo_image.add_header("Content-ID", "<logo_utpl>")
        logo_image.add_header("Content-Disposition", "inline", filename="logo-utpl.png")
        email.attach(logo_image)

    return email


@login_required
@require_POST
def siniestro_email_enviar(request, pk, tipo):
    """
    Procesa el envío del email con el contenido editado por el usuario.
    Usa la plantilla HTML base_notificacion.html para emails con formato profesional.
    """
    from django.conf import settings
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string

    siniestro = get_object_or_404(Siniestro, pk=pk)

    destinatario = request.POST.get("destinatario", "")
    cc = request.POST.get("cc", "")
    asunto = request.POST.get("asunto", "")
    contenido = request.POST.get("contenido", "")
    notas_adicionales = request.POST.get("notas_adicionales", "")

    if not destinatario or not asunto or not contenido:
        messages.error(request, "Faltan campos obligatorios en el email.")
        return redirect("siniestro_detalle", pk=pk)

    # Agregar notas adicionales si existen
    if notas_adicionales:
        if "Atentamente," in contenido:
            contenido = contenido.replace(
                "Atentamente,", f"NOTAS ADICIONALES\n=================\n{notas_adicionales}\n\nAtentamente,"
            )
        else:
            contenido += f"\n\nNOTAS ADICIONALES\n=================\n{notas_adicionales}"

    try:
        cc_list = [e.strip() for e in cc.split(",") if e.strip()] if cc else []

        # Generar HTML usando la plantilla
        contenido_html = _generar_html_email(siniestro, tipo, asunto, contenido, notas_adicionales)

        # Crear email con alternativa HTML y logo embebido
        email = _crear_email_con_logo(
            asunto=asunto,
            contenido_texto=contenido,
            contenido_html=contenido_html,
            destinatario=destinatario,
            cc_list=cc_list,
        )

        # Si es aseguradora, adjuntar archivos
        if tipo == "aseguradora":
            for adjunto in siniestro.adjuntos.all():
                if adjunto.archivo:
                    try:
                        email.attach_file(adjunto.archivo.path)
                    except Exception:
                        pass  # Ignorar adjuntos que no se puedan cargar

        # Enviar
        email.send(fail_silently=False)

        # Actualizar estado según tipo (usando métodos de transición)
        if tipo == "broker":
            # Transición: registrado → notificado_broker
            if siniestro.puede_notificar_broker:
                siniestro.notificar_broker()
                mensaje_exito = "Broker notificado. Esperando su respuesta para continuar."
            else:
                siniestro.fecha_notificacion_broker = timezone.now()
                siniestro.save(update_fields=["fecha_notificacion_broker"])
                mensaje_exito = "Notificación enviada al broker."

            # También notificar al custodio que el proceso inició
            _notificar_custodio_inicio_proceso(siniestro)
        elif tipo == "aseguradora":
            # Transición: documentacion_lista → enviado_aseguradora
            if siniestro.puede_enviar_aseguradora:
                siniestro.enviar_a_aseguradora()
                mensaje_exito = "Documentos enviados a la aseguradora. Esperando recibo de indemnización."
            else:
                siniestro.estado = "enviado_aseguradora"
                siniestro.fecha_envio_aseguradora = timezone.now().date()
                siniestro.save(update_fields=["estado", "fecha_envio_aseguradora"])
                mensaje_exito = "Siniestro enviado a la aseguradora."
        elif tipo == "responsable":
            siniestro.fecha_notificacion_responsable = timezone.now().date()
            siniestro.save(update_fields=["fecha_notificacion_responsable"])
            mensaje_exito = "Notificación enviada al responsable."
        else:
            mensaje_exito = "Email enviado exitosamente."

        # Registrar en el historial de notificaciones
        NotificacionEmail.objects.create(
            tipo=f"siniestro_{tipo}",
            destinatario=destinatario,
            cc=cc,
            asunto=asunto,
            contenido=contenido,
            contenido_html=contenido_html,
            siniestro=siniestro,
            estado="enviado",
            fecha_envio=timezone.now(),
            creado_por=request.user,
        )

        messages.success(request, mensaje_exito)

    except Exception as e:
        messages.error(request, f"Error al enviar el email: {str(e)}")

        # Registrar el error
        NotificacionEmail.objects.create(
            tipo=f"siniestro_{tipo}",
            destinatario=destinatario,
            cc=cc,
            asunto=asunto,
            contenido=contenido,
            siniestro=siniestro,
            estado="error",
            error_mensaje=str(e),
            creado_por=request.user,
        )

    return redirect("siniestro_detalle", pk=pk)


@login_required
@require_POST
def siniestro_notificar_broker(request, pk):
    """Notifica el siniestro al broker (versión legacy, redirige al modal)"""
    siniestro = get_object_or_404(Siniestro, pk=pk)

    try:
        NotificacionesService.notificar_siniestro_a_broker(siniestro, request.user)
        messages.success(request, "Notificación enviada al broker exitosamente.")
    except Exception as e:
        messages.error(request, f"Error al enviar notificación: {str(e)}")

    return redirect("siniestro_detalle", pk=pk)


@login_required
@require_POST
def siniestro_marcar_checklist(request, siniestro_pk, item_pk):
    """Marca un item de checklist como completado"""
    checklist_item = get_object_or_404(ChecklistSiniestro, pk=item_pk, siniestro_id=siniestro_pk)

    checklist_item.marcar_completado(request.user)
    messages.success(request, f'Item "{checklist_item.config_item.nombre}" marcado como completado.')

    return redirect("siniestro_detalle", pk=siniestro_pk)


@login_required
def siniestro_descargar_carta(request, pk):
    """Descarga la carta formal del siniestro"""
    siniestro = get_object_or_404(Siniestro, pk=pk)
    return DocumentosService.descargar_carta_siniestro(siniestro)


@login_required
def siniestro_descargar_recibo(request, pk):
    """Descarga el recibo de indemnización"""
    siniestro = get_object_or_404(Siniestro, pk=pk)
    return DocumentosService.descargar_recibo_indemnizacion(siniestro)


@login_required
@require_POST
def adjunto_firmar(request, pk):
    """Aplica firma electrónica a un adjunto"""
    adjunto = get_object_or_404(AdjuntoSiniestro, pk=pk)

    try:
        ip = request.META.get("REMOTE_ADDR")
        DocumentosService.aplicar_firma_electronica(adjunto, request.user, ip)
        messages.success(request, "Documento firmado exitosamente.")
    except Exception as e:
        messages.error(request, f"Error al firmar documento: {str(e)}")

    return redirect("siniestro_detalle", pk=adjunto.siniestro.pk)


@login_required
@require_POST
def siniestro_subir_adjunto(request, pk):
    """Sube un adjunto al siniestro y opcionalmente marca el checklist"""
    siniestro = get_object_or_404(Siniestro, pk=pk)

    archivo = request.FILES.get("archivo")
    nombre = request.POST.get("nombre", "")
    tipo_adjunto = request.POST.get("tipo_adjunto", "otro")
    checklist_item_id = request.POST.get("checklist_item_id")
    marcar_completado = request.POST.get("marcar_completado") == "on"

    if not archivo:
        messages.error(request, "Debes seleccionar un archivo.")
        return redirect("siniestro_detalle", pk=pk)

    # Validar tamaño (máximo 10MB)
    if archivo.size > 10 * 1024 * 1024:
        messages.error(request, "El archivo no puede superar los 10MB.")
        return redirect("siniestro_detalle", pk=pk)

    # Crear el adjunto
    checklist_item = None
    if checklist_item_id:
        checklist_item = ChecklistSiniestro.objects.filter(pk=checklist_item_id, siniestro=siniestro).first()

    _ = AdjuntoSiniestro.objects.create(
        siniestro=siniestro,
        nombre=nombre or archivo.name,
        archivo=archivo,
        tipo_adjunto=tipo_adjunto,
        checklist_item=checklist_item,
        subido_por=request.user,
    )

    # Si tiene checklist_item y se pide marcar como completado
    if checklist_item and marcar_completado:
        checklist_item.marcar_completado(request.user)
        messages.success(request, f'Documento subido y "{checklist_item.config_item.nombre}" marcado como completado.')
    else:
        messages.success(request, "Documento subido exitosamente.")

    return redirect("siniestro_detalle", pk=pk)


@login_required
@require_POST
def siniestro_enviar_aseguradora(request, pk):
    """Envía el siniestro con todos sus documentos a la aseguradora"""
    from django.conf import settings
    from django.core.mail import EmailMessage

    siniestro = get_object_or_404(Siniestro, pk=pk)

    # Verificar que el checklist de obligatorios esté completo
    checklist_items = siniestro.checklist_items.select_related("config_item").all()
    obligatorios_pendientes = [
        item for item in checklist_items if item.config_item.es_obligatorio and not item.completado
    ]

    if obligatorios_pendientes:
        nombres = ", ".join([item.config_item.nombre for item in obligatorios_pendientes])
        messages.error(request, f"Faltan documentos obligatorios: {nombres}")
        return redirect("siniestro_detalle", pk=pk)

    # Obtener adjuntos
    adjuntos = siniestro.adjuntos.all()

    # Preparar email
    asunto = f"Siniestro {siniestro.numero_siniestro} - {siniestro.poliza.numero_poliza}"

    mensaje = f"""
Estimados señores,

Por medio de la presente, nos permitimos enviar la documentación correspondiente al siniestro:

Número de Siniestro: {siniestro.numero_siniestro}
Póliza: {siniestro.poliza.numero_poliza}
Tipo de Siniestro: {siniestro.tipo_siniestro.get_nombre_display() if siniestro.tipo_siniestro else 'N/A'}
Fecha del Siniestro: {siniestro.fecha_siniestro.strftime('%d/%m/%Y')}
Bien Afectado: {siniestro.bien_nombre}
Monto Estimado: ${siniestro.monto_estimado:,.2f}

Causa: {siniestro.causa}

Se adjuntan los siguientes documentos:
"""

    for adjunto in adjuntos:
        mensaje += f"- {adjunto.nombre} ({adjunto.get_tipo_adjunto_display()})\n"

    mensaje += """
Quedamos atentos a su respuesta.

Atentamente,
Gestión de Seguros
"""

    try:
        # Email de destino configurado
        email_destino = "renataxdalej@gmail.com"

        # Lista de destinatarios CC (usuario actual y broker si existe)
        cc_list = []
        if request.user.email:
            cc_list.append(request.user.email)
        if siniestro.poliza.corredor_seguros and siniestro.poliza.corredor_seguros.email:
            cc_list.append(siniestro.poliza.corredor_seguros.email)

        # Crear email con adjuntos
        email = EmailMessage(
            subject=asunto,
            body=mensaje,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_destino],
            cc=cc_list if cc_list else None,
        )

        # Adjuntar archivos
        for adjunto in adjuntos:
            if adjunto.archivo:
                email.attach(adjunto.nombre, adjunto.archivo.read(), "application/octet-stream")

        email.send()

        # Actualizar estado del siniestro
        siniestro.estado = "enviado_aseguradora"
        siniestro.fecha_envio_aseguradora = timezone.now()
        siniestro.save()

        messages.success(request, "Siniestro enviado a la aseguradora exitosamente.")
    except Exception as e:
        messages.error(request, f"Error al enviar el siniestro: {str(e)}")

    return redirect("siniestro_detalle", pk=pk)


# =============================================================================
# API ENDPOINTS ADICIONALES
# =============================================================================


@login_required
@require_GET
def api_subtipos_ramo(request):
    """API para obtener subgrupos de un grupo de ramo"""
    ramo_id = request.GET.get("ramo_id")
    grupo_ramo_id = request.GET.get("grupo_ramo_id") or ramo_id  # Compatibilidad
    if not grupo_ramo_id:
        return JsonResponse({"subtipos": [], "subgrupos": []})

    subgrupos = (
        SubgrupoRamo.objects.filter(grupo_ramo_id=grupo_ramo_id, activo=True)
        .values("id", "codigo", "nombre")
        .order_by("orden", "nombre")
    )

    resultado = list(subgrupos)
    return JsonResponse({"subtipos": resultado, "subgrupos": resultado})


@login_required
@require_GET
def api_corredores_por_compania(request):
    """API para obtener corredores de una compañía aseguradora"""
    compania_id = request.GET.get("compania_id")
    if not compania_id:
        return JsonResponse({"corredores": []})

    corredores = (
        CorredorSeguros.objects.filter(compania_aseguradora_id=compania_id, activo=True)
        .values("id", "nombre", "email")
        .order_by("nombre")
    )

    return JsonResponse({"corredores": list(corredores)})


@login_required
@require_GET
def api_calcular_desglose_ramo(request):
    """API para calcular el desglose financiero de un ramo"""
    try:
        prima = Decimal(request.GET.get("prima", "0"))
        es_gran_contribuyente = request.GET.get("gran_contribuyente", "false") == "true"
    except (ValueError, TypeError):
        return JsonResponse({"error": "Valores inválidos"}, status=400)

    # Calcular contribuciones sobre la prima
    contrib_super = prima * Decimal("0.035")
    seguro_camp = prima * Decimal("0.005")

    # Calcular derechos de emisión en backend (tabla escalonada)
    emision = DetallePolizaRamo.calcular_derechos_emision(prima)

    # Calcular base imponible, IVA y total
    base_imponible = prima + contrib_super + seguro_camp + emision
    iva = base_imponible * Decimal("0.15")
    total_facturado = base_imponible + iva

    if es_gran_contribuyente:
        ret_prima = prima * Decimal("0.01")
        ret_iva = iva
    else:
        ret_prima = Decimal("0")
        ret_iva = Decimal("0")

    valor_por_pagar = total_facturado - ret_prima - ret_iva

    return JsonResponse(
        {
            "contribucion_superintendencia": float(contrib_super),
            "seguro_campesino": float(seguro_camp),
            "base_imponible": float(base_imponible),
            "iva": float(iva),
            "total_facturado": float(total_facturado),
            "retencion_prima": float(ret_prima),
            "retencion_iva": float(ret_iva),
            "valor_por_pagar": float(valor_por_pagar),
        }
    )


@login_required
@require_GET
def api_reporte_siniestralidad(request):
    """API para obtener datos de siniestralidad"""
    from datetime import datetime

    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    compania_id = request.GET.get("compania_id")

    if fecha_desde:
        fecha_desde = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    if fecha_hasta:
        fecha_hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()

    datos = ReportesAvanzadosService.calcular_siniestralidad(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        compania_id=compania_id,
    )

    return JsonResponse(datos)


@login_required
def inicializar_ramos_predefinidos(request):
    """Inicializa los ramos predefinidos del sistema"""
    Ramo.crear_ramos_predefinidos()
    messages.success(request, "Ramos predefinidos creados exitosamente.")
    return redirect("ramos_lista")


# ==========================================================================
# VISTAS DE FACTURA - CRUD Completo
# ==========================================================================


class FacturaCreateView(LoginRequiredMixin, CreateView):
    """Vista para crear una nueva factura"""

    model = Factura
    form_class = FacturaForm
    template_name = "app/facturas/crear.html"
    success_url = reverse_lazy("facturas_lista")

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, "Factura creada exitosamente.")
        return super().form_valid(form)


class FacturaDetailView(LoginRequiredMixin, DetailView):
    """Vista de detalle de una factura"""

    model = Factura
    template_name = "app/facturas/detalle.html"
    context_object_name = "factura"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["pagos"] = self.object.pagos.all().order_by("-fecha_pago")
        context["documentos"] = self.object.documentos.all()
        return context


class FacturaUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para editar una factura"""

    model = Factura
    form_class = FacturaForm
    template_name = "app/facturas/editar.html"

    def get_success_url(self):
        return reverse("factura_detalle", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Factura actualizada exitosamente.")
        return super().form_valid(form)


# ==========================================================================
# VISTAS DE DOCUMENTO - CRUD Completo
# ==========================================================================


class DocumentoCreateView(LoginRequiredMixin, CreateView):
    """Vista para crear un nuevo documento"""

    model = Documento
    form_class = DocumentoForm
    template_name = "app/documentos/crear.html"
    success_url = reverse_lazy("documentos_lista")

    def form_valid(self, form):
        form.instance.subido_por = self.request.user
        messages.success(self.request, "Documento subido exitosamente.")
        return super().form_valid(form)


class DocumentoUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para editar un documento"""

    model = Documento
    form_class = DocumentoForm
    template_name = "app/documentos/editar.html"

    def get_success_url(self):
        return reverse("documento_ver", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Documento actualizado exitosamente.")
        return super().form_valid(form)


# ==========================================================================
# VISTAS DE PAGO - CRUD Completo
# ==========================================================================


class PagoCreateView(LoginRequiredMixin, CreateView):
    """Vista para registrar un nuevo pago"""

    model = Pago
    form_class = PagoForm
    template_name = "app/pagos/crear.html"
    success_url = reverse_lazy("facturas_lista")

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        messages.success(self.request, "Pago registrado exitosamente.")
        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        factura_id = self.request.GET.get("factura")
        if factura_id:
            initial["factura"] = factura_id
        return initial


class PagoDetailView(LoginRequiredMixin, DetailView):
    """Vista de detalle de un pago"""

    model = Pago
    template_name = "app/pagos/detalle.html"
    context_object_name = "pago"


class PagoUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para editar un pago"""

    model = Pago
    form_class = PagoForm
    template_name = "app/pagos/editar.html"

    def get_success_url(self):
        return reverse("factura_detalle", kwargs={"pk": self.object.factura.pk})

    def form_valid(self, form):
        messages.success(self.request, "Pago actualizado exitosamente.")
        return super().form_valid(form)


# ==========================================================================
# VISTAS DE SINIESTROS POR EMAIL
# ==========================================================================


@login_required
def siniestros_email_pendientes(request):
    """
    Vista para listar y gestionar los siniestros reportados por email.
    Permite filtrar por estado y completar la información faltante.
    """
    queryset = SiniestroEmail.objects.select_related(
        "activo_encontrado", "siniestro_creado", "responsable_encontrado"
    ).order_by("-fecha_recepcion")

    # Filtros
    query = request.GET.get("q", "").strip()
    if query:
        queryset = queryset.filter(
            Q(serie__icontains=query)
            | Q(responsable_nombre__icontains=query)
            | Q(email_subject__icontains=query)
            | Q(marca__icontains=query)
            | Q(modelo__icontains=query)
        )

    estado = request.GET.get("estado")
    if estado:
        queryset = queryset.filter(estado_procesamiento=estado)

    # Paginación
    paginator = Paginator(queryset, 15)
    page = request.GET.get("page", 1)
    pendientes = paginator.get_page(page)

    # Estadísticas
    stats = {
        "total": SiniestroEmail.objects.count(),
        "pendientes": SiniestroEmail.objects.filter(estado_procesamiento="pendiente").count(),
        "procesados": SiniestroEmail.objects.filter(estado_procesamiento="procesado").count(),
        "errores": SiniestroEmail.objects.filter(estado_procesamiento="error").count(),
    }

    # Datos para el modal de completar
    polizas = Poliza.objects.filter(estado="vigente").select_related("compania_aseguradora")
    tipos_siniestro = TipoSiniestro.objects.filter(activo=True)
    responsables = ResponsableCustodio.objects.filter(activo=True).order_by("nombre")

    context = {
        "pendientes": pendientes,
        "stats": stats,
        "query": query,
        "estado_filtro": estado,
        "polizas": polizas,
        "tipos_siniestro": tipos_siniestro,
        "responsables": responsables,
    }

    return render(request, "app/siniestros/email_pendientes.html", context)


@login_required
def siniestro_email_procesar_auto(request, pk):
    """
    Intenta crear el siniestro automáticamente buscando el bien por número de serie.
    Si no encuentra el bien, redirige al formulario manual con mensaje.
    """
    from datetime import datetime

    siniestro_email = get_object_or_404(SiniestroEmail, pk=pk)

    # Verificar que esté pendiente
    if siniestro_email.estado_procesamiento != "pendiente":
        messages.warning(request, "Este registro ya fue procesado anteriormente.")
        return redirect("siniestros_email_pendientes")

    # Buscar el bien por número de serie (modelo unificado BienAsegurado)
    bien = BienAsegurado.objects.filter(serie__iexact=siniestro_email.serie.strip()).select_related("poliza").first()

    if not bien:
        messages.warning(
            request,
            f'No se encontró un bien registrado con serie "{siniestro_email.serie}". '
            "Por favor complete los datos manualmente.",
        )
        return redirect("siniestros_email_pendientes")

    if not bien.poliza:
        messages.warning(
            request, f'El bien "{bien.nombre}" no tiene póliza asignada. ' "Por favor complete los datos manualmente."
        )
        return redirect("siniestros_email_pendientes")

    # Buscar tipo de siniestro "daño" por defecto
    tipo_siniestro = TipoSiniestro.objects.filter(nombre="daño").first()
    if not tipo_siniestro:
        tipo_siniestro = TipoSiniestro.objects.filter(activo=True).first()

    if not tipo_siniestro:
        messages.error(request, "No hay tipos de siniestro configurados en el sistema.")
        return redirect("siniestros_email_pendientes")

    # Parsear fecha del reporte
    fecha_siniestro = timezone.now()
    if siniestro_email.fecha_reporte:
        try:
            fecha_siniestro = timezone.make_aware(datetime.strptime(siniestro_email.fecha_reporte.strip(), "%d/%m/%Y"))
        except ValueError:
            pass

    # Generar número de siniestro
    from django.db.models import Max

    ultimo = Siniestro.objects.aggregate(Max("id"))["id__max"] or 0
    numero_siniestro = f"SIN-{timezone.now().year}-{str(ultimo + 1).zfill(5)}"

    try:
        with transaction.atomic():
            # Crear el siniestro usando datos del bien
            siniestro = Siniestro.objects.create(
                poliza=bien.poliza,
                bien_asegurado=bien,
                numero_siniestro=numero_siniestro,
                tipo_siniestro=tipo_siniestro,
                fecha_siniestro=fecha_siniestro,
                bien_nombre=bien.nombre,
                bien_modelo=bien.modelo or "",
                bien_serie=bien.serie or "",
                bien_marca=bien.marca or "",
                bien_codigo_activo=bien.codigo_activo or "",
                responsable_custodio=bien.responsable_custodio,
                ubicacion=bien.ubicacion or "",
                causa=siniestro_email.causa,
                descripcion_detallada=siniestro_email.problema,
                monto_estimado=bien.valor_actual or bien.valor_asegurado,
                estado="registrado",
            )

            # Crear checklist inicial
            items_config = ChecklistSiniestroConfig.objects.filter(tipo_siniestro=tipo_siniestro, activo=True)
            for config in items_config:
                ChecklistSiniestro.objects.create(siniestro=siniestro, config_item=config, completado=False)

            # Actualizar el registro de email
            siniestro_email.activo_encontrado = bien
            siniestro_email.siniestro_creado = siniestro
            siniestro_email.responsable_encontrado = bien.responsable_custodio
            siniestro_email.estado_procesamiento = "procesado"
            siniestro_email.fecha_procesamiento = timezone.now()
            siniestro_email.procesado_por = request.user
            siniestro_email.mensaje_procesamiento = (
                f"Siniestro {numero_siniestro} creado automáticamente. " f"Bien: {bien.codigo_activo} - {bien.nombre}"
            )
            siniestro_email.save()

        messages.success(
            request,
            f"✅ Siniestro {numero_siniestro} creado automáticamente. "
            f"Bien: {bien.nombre} | Póliza: {bien.poliza.numero_poliza}",
        )
        return redirect("siniestro_detalle", pk=siniestro.pk)

    except Exception as e:
        messages.error(request, f"Error al crear el siniestro: {str(e)}")
        return redirect("siniestros_email_pendientes")


@login_required
@require_POST
def siniestro_email_completar(request, pk):
    """
    Completa la información de un siniestro recibido por email manualmente.

    La vista solo:
    1. Extrae datos del request
    2. Delega al servicio
    3. Maneja la respuesta HTTP

    Toda la lógica de negocio está en SiniestroService.crear_desde_email()
    """
    from app.services.siniestro import SiniestroService

    siniestro_email = get_object_or_404(SiniestroEmail, pk=pk)

    # Verificar estado (validación de flujo HTTP, no de negocio)
    if siniestro_email.estado_procesamiento != "pendiente":
        messages.error(request, "Este registro ya fue procesado anteriormente.")
        return redirect("siniestros_email_pendientes")

    # Extraer datos del request (responsabilidad de la vista)
    poliza_id = request.POST.get("poliza")
    tipo_siniestro_id = request.POST.get("tipo_siniestro")
    ubicacion = request.POST.get("ubicacion", "").strip()
    monto_estimado = request.POST.get("monto_estimado")
    responsable_id = request.POST.get("responsable_custodio")

    # Validación básica de campos requeridos (nivel HTTP)
    if not all([poliza_id, tipo_siniestro_id, ubicacion, monto_estimado]):
        messages.error(request, "Por favor complete todos los campos obligatorios.")
        return redirect("siniestros_email_pendientes")

    try:
        # Obtener entidades relacionadas
        poliza = Poliza.objects.get(pk=poliza_id)
        tipo_siniestro = TipoSiniestro.objects.get(pk=tipo_siniestro_id)
        monto = Decimal(monto_estimado)
        responsable = ResponsableCustodio.objects.get(pk=responsable_id) if responsable_id else None

        # Delegar TODA la lógica de negocio al servicio
        resultado = SiniestroService.crear_desde_email(
            siniestro_email=siniestro_email,
            poliza=poliza,
            tipo_siniestro=tipo_siniestro,
            ubicacion=ubicacion,
            monto_estimado=monto,
            responsable=responsable,
            fecha_reporte_str=siniestro_email.fecha_reporte,
            usuario=request.user,
        )

        # Manejar resultado
        if resultado.exitoso:
            messages.success(request, resultado.mensaje)
            return redirect("siniestro_detalle", pk=resultado.objeto.pk)
        else:
            for campo, error in resultado.errores.items():
                messages.error(request, f"{campo}: {error}")
            return redirect("siniestros_email_pendientes")

    except Poliza.DoesNotExist:
        messages.error(request, "La póliza seleccionada no existe.")
    except TipoSiniestro.DoesNotExist:
        messages.error(request, "El tipo de siniestro seleccionado no existe.")
    except Exception as e:
        messages.error(request, f"Error al crear el siniestro: {str(e)}")

    return redirect("siniestros_email_pendientes")


@login_required
def siniestros_email_count(request):
    """API para obtener el conteo de siniestros email pendientes."""
    count = SiniestroEmail.objects.filter(estado_procesamiento="pendiente").count()
    return JsonResponse({"count": count})


# ==============================================================================
# ESCANEO RÁPIDO DE INBOX IMAP
# ==============================================================================


@login_required
@require_POST
def escanear_reportes_custodio(request):
    """
    Vista para escanear manualmente el inbox buscando reportes de siniestros de custodios.
    """
    try:
        from app.services.email.reader import procesar_y_guardar_correos

        resultado = procesar_y_guardar_correos(unseen_only=True, mark_as_read=True)

        return JsonResponse(
            {
                "success": True,
                "encontrados": resultado.get("total_procesados", 0),
                "siniestros_creados": resultado.get("siniestros_creados", 0),
                "mensaje": f"Procesados {resultado.get('total_procesados', 0)} correos",
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
@require_POST
def escanear_inbox(request, pk):
    """
    Vista para escanear rápidamente el inbox IMAP según el estado del siniestro.
    """
    from django.http import JsonResponse

    siniestro = get_object_or_404(Siniestro, pk=pk)
    tipo = request.GET.get("tipo", "broker")

    try:
        encontrado = False
        mensaje = ""

        if tipo == "broker":
            # Escanear respuestas del broker
            from app.services.email.broker_reader import procesar_respuestas_broker

            resultado = procesar_respuestas_broker()
            encontrado = resultado.get("vinculadas", 0) > 0
            mensaje = f"Encontradas {resultado.get('vinculadas', 0)} respuestas"

        elif tipo == "recibos":
            # Escanear recibos de indemnización
            from app.services.email.recibos_reader import procesar_recibos_indemnizacion

            resultado = procesar_recibos_indemnizacion()
            encontrado = resultado.get("vinculados", 0) > 0
            mensaje = f"Procesados {resultado.get('procesados', 0)}, vinculados {resultado.get('vinculados', 0)}"

        # Refrescar el siniestro para ver si cambió de estado
        siniestro.refresh_from_db()

        return JsonResponse(
            {"success": True, "encontrado": encontrado, "mensaje": mensaje, "nuevo_estado": siniestro.estado}
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ==============================================================================
# VISTAS DEL FLUJO DE INDEMNIZACIÓN
# ==============================================================================


@login_required
def siniestro_firmar_recibo(request, pk):
    """
    Vista para firmar el recibo de indemnización.
    El usuario sube el PDF firmado y confirma que está de acuerdo.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if not siniestro.puede_firmar_recibo:
        messages.error(request, "El siniestro no está en estado válido para firmar el recibo.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        recibo_firmado = request.FILES.get("recibo_firmado")
        conforme = request.POST.get("conforme") == "on"

        if not recibo_firmado:
            messages.error(request, "Debe adjuntar el recibo firmado.")
            return redirect("siniestro_detalle", pk=pk)

        if not conforme:
            messages.error(request, "Debe confirmar que está de acuerdo con la indemnización.")
            return redirect("siniestro_detalle", pk=pk)

        # Usar método de transición
        filename = f"recibo_firmado_{siniestro.numero_siniestro}_{timezone.now().strftime('%Y%m%d')}.pdf"
        from django.core.files.base import ContentFile

        archivo_content = ContentFile(recibo_firmado.read())
        archivo_content.name = filename
        siniestro.firmar_recibo(archivo_content)

        messages.success(request, "Recibo firmado. Ahora puede enviarlo a la aseguradora para iniciar el plazo de 72h.")
        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/firmar_recibo.html",
        {
            "siniestro": siniestro,
        },
    )


@login_required
def siniestro_disputar(request, pk):
    """
    Vista para iniciar una disputa sobre el recibo de indemnización.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if not siniestro.puede_disputar:
        messages.error(request, "El siniestro no está en estado válido para disputar.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        motivo = request.POST.get("motivo", "").strip()

        if not motivo:
            messages.error(request, "Debe indicar el motivo de la disputa.")
            return redirect("siniestro_detalle", pk=pk)

        siniestro.registrar_disputa(motivo)

        # Enviar email a la aseguradora notificando la disputa
        # TODO: Implementar envío de email

        messages.success(request, "Disputa registrada correctamente.")
        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/disputar.html",
        {
            "siniestro": siniestro,
        },
    )


@login_required
def siniestro_resolver_disputa(request, pk):
    """
    Vista para resolver una disputa sobre el recibo.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if siniestro.estado != "en_disputa":
        messages.error(request, "El siniestro no está en disputa.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        resolucion = request.POST.get("resolucion", "").strip()
        nuevo_recibo = request.FILES.get("nuevo_recibo")

        if not resolucion:
            messages.error(request, "Debe indicar cómo se resolvió la disputa.")
            return redirect("siniestro_detalle", pk=pk)

        # Resolver disputa
        siniestro.resolver_disputa(resolucion)

        # Si hay nuevo recibo, actualizar
        if nuevo_recibo:
            filename = f"recibo_{siniestro.numero_siniestro}_{timezone.now().strftime('%Y%m%d')}.pdf"
            siniestro.recibo_indemnizacion.save(filename, nuevo_recibo, save=False)
            siniestro.fecha_recibo_recibido = timezone.now()
            siniestro.save(update_fields=["recibo_indemnizacion", "fecha_recibo_recibido"])

        messages.success(request, 'Disputa resuelta. El siniestro vuelve a estado "Recibo Recibido".')
        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/resolver_disputa.html",
        {
            "siniestro": siniestro,
        },
    )


@login_required
def siniestro_enviar_liquidacion(request, pk):
    """
    Vista para enviar el recibo firmado a la aseguradora e iniciar el contador de 72h.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if not siniestro.puede_enviar_liquidacion:
        messages.error(request, "El siniestro no está listo para enviar a liquidación.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        # Iniciar el proceso de liquidación (calcula fecha límite 72h hábiles)
        siniestro.iniciar_liquidacion()

        # Enviar email a la aseguradora con el recibo firmado
        poliza = siniestro.get_poliza()
        if poliza and poliza.compania_aseguradora and poliza.compania_aseguradora.email:
            try:
                from django.template.loader import render_to_string

                from app.models import ConfiguracionSistema

                # Obtener datos del firmante
                firmante_nombre = ConfiguracionSistema.get_config("FIRMANTE_CARTA_NOMBRE", "Gestión de Seguros")
                firmante_cargo = ConfiguracionSistema.get_config("FIRMANTE_CARTA_CARGO", "Seguros / Inventarios")
                firmante_departamento = ConfiguracionSistema.get_config("FIRMANTE_DEPARTAMENTO", "")
                firmante_telefono = ConfiguracionSistema.get_config("FIRMANTE_TELEFONO", "")
                firmante_email_config = ConfiguracionSistema.get_config("FIRMANTE_EMAIL", "")

                bien = siniestro.bien_asegurado
                bien_nombre = bien.nombre if bien else siniestro.get_nombre_bien()

                asunto = f"Recibo de Indemnización Firmado - Siniestro {siniestro.numero_siniestro}"

                # Contexto para la plantilla
                context = {
                    "numero_siniestro": siniestro.numero_siniestro,
                    "numero_poliza": poliza.numero_poliza,
                    "bien_nombre": bien_nombre,
                    "fecha_firma": (
                        siniestro.fecha_firma_indemnizacion.strftime("%d/%m/%Y")
                        if siniestro.fecha_firma_indemnizacion
                        else timezone.now().strftime("%d/%m/%Y")
                    ),
                    "fecha_limite": (
                        siniestro.fecha_limite_liquidacion.strftime("%d/%m/%Y %H:%M")
                        if siniestro.fecha_limite_liquidacion
                        else "N/A"
                    ),
                    "firmante_nombre": firmante_nombre,
                    "firmante_cargo": firmante_cargo,
                    "firmante_departamento": firmante_departamento,
                    "firmante_telefono": firmante_telefono,
                    "firmante_email": firmante_email_config,
                    "timestamp": timezone.now().isoformat(),
                }

                # Renderizar plantilla HTML
                contenido_html = render_to_string("emails/recibo_firmado.html", context)

                # Versión texto plano
                contenido_texto = f"""Estimados,

Por medio de la presente, adjunto el recibo de indemnización debidamente firmado correspondiente al siniestro No. {siniestro.numero_siniestro}.

Solicitamos amablemente proceder con la liquidación dentro del plazo establecido de 72 horas hábiles.

Fecha límite de liquidación: {context['fecha_limite']}

Quedamos atentos a su confirmación.

Saludos cordiales,
{firmante_nombre}
{firmante_cargo}
"""

                # Usar función de crear email con logo
                email = _crear_email_con_logo(
                    asunto=asunto,
                    contenido_texto=contenido_texto,
                    contenido_html=contenido_html,
                    destinatario=poliza.compania_aseguradora.email,
                )

                # Adjuntar recibo firmado
                if siniestro.recibo_firmado:
                    email.attach_file(siniestro.recibo_firmado.path)

                email.send()
                messages.success(
                    request, f"Recibo enviado a {poliza.compania_aseguradora.email}. Plazo de 72h hábiles iniciado."
                )
            except Exception as e:
                messages.warning(request, f"Proceso iniciado pero hubo error enviando email: {str(e)}")
        else:
            messages.success(request, "Proceso de liquidación iniciado. Plazo de 72h hábiles comenzó.")

        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/enviar_liquidacion.html",
        {
            "siniestro": siniestro,
        },
    )


@login_required
def siniestro_registrar_liquidacion(request, pk):
    """
    Vista para registrar la liquidación del siniestro.
    Incluye validación de diferencia de montos.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if not siniestro.puede_registrar_liquidacion:
        messages.error(request, "El siniestro no está en estado válido para registrar liquidación.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        try:
            from datetime import datetime

            monto = Decimal(request.POST.get("monto", "0"))
            comprobante = request.POST.get("comprobante", "").strip()
            fecha_pago_str = request.POST.get("fecha_pago", "").strip()
            motivo_diferencia = request.POST.get("motivo_diferencia", "").strip()
            confirmar_diferencia = request.POST.get("confirmar_diferencia") == "on"

            # Parsear fecha de pago
            fecha_pago = None
            if fecha_pago_str:
                try:
                    fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            if monto <= 0:
                messages.error(request, "El monto debe ser mayor a cero.")
                return redirect("siniestro_detalle", pk=pk)

            if not comprobante:
                messages.error(request, "Debe indicar el número de comprobante.")
                return redirect("siniestro_detalle", pk=pk)

            # Verificar diferencia de monto
            monto_esperado = siniestro.monto_a_indemnizar or Decimal("0")
            diferencia = abs(monto - monto_esperado)
            hay_diferencia = diferencia > Decimal("0.01")

            if hay_diferencia and not confirmar_diferencia:
                messages.error(request, "Hay diferencia en el monto. Debe confirmar y explicar la diferencia.")
                return render(
                    request,
                    "app/siniestros/registrar_liquidacion.html",
                    {
                        "siniestro": siniestro,
                        "monto": monto,
                        "comprobante": comprobante,
                        "fecha_pago": fecha_pago_str,
                        "diferencia": monto - monto_esperado,
                        "mostrar_diferencia": True,
                    },
                )

            # Registrar liquidación con fecha de pago
            siniestro.registrar_liquidacion(monto, comprobante, motivo_diferencia, fecha_pago)

            messages.success(request, "Liquidación registrada correctamente.")
            return redirect("siniestro_detalle", pk=pk)

        except (ValueError, TypeError):
            messages.error(request, "Monto inválido.")
            return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/registrar_liquidacion.html",
        {
            "siniestro": siniestro,
            "monto_esperado": siniestro.monto_a_indemnizar,
            "fecha_pago": timezone.now().date().isoformat(),  # Fecha de hoy como default
        },
    )


@login_required
def siniestro_cerrar(request, pk):
    """
    Vista para cerrar el siniestro después de liquidación.
    Envía notificaciones al responsable y gerencia.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if siniestro.estado != "liquidado":
        messages.error(request, "El siniestro debe estar liquidado para cerrarlo.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        # Cerrar siniestro
        siniestro.cerrar_siniestro()

        # Enviar notificaciones
        _enviar_notificaciones_cierre(siniestro)

        messages.success(request, "Siniestro cerrado. Se enviaron notificaciones al responsable y gerencia.")
        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/cerrar.html",
        {
            "siniestro": siniestro,
        },
    )


@login_required
@login_required
def siniestro_marcar_docs_listos(request, pk):
    """
    Vista para marcar la documentación como lista.
    Cambia el estado de 'documentacion_pendiente' a 'documentacion_lista'.
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if siniestro.estado != "documentacion_pendiente":
        messages.error(request, "El siniestro no está en estado de documentación pendiente.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        siniestro.estado = "documentacion_lista"
        siniestro.save(update_fields=["estado"])
        messages.success(request, "Documentación marcada como lista. Ahora puede enviar a la aseguradora.")
        return redirect("siniestro_detalle", pk=pk)

    # GET: mostrar confirmación simple
    return render(
        request,
        "app/siniestros/confirmar_docs_listos.html",
        {
            "siniestro": siniestro,
        },
    )


def siniestro_rechazar(request, pk):
    """
    Vista para rechazar un siniestro (cierre sin pago).
    """
    siniestro = get_object_or_404(Siniestro, pk=pk)

    if siniestro.estado in ["cerrado", "rechazado", "liquidado"]:
        messages.error(request, "El siniestro ya está cerrado o liquidado.")
        return redirect("siniestro_detalle", pk=pk)

    if request.method == "POST":
        motivo = request.POST.get("motivo", "").strip()

        if not motivo:
            messages.error(request, "Debe indicar el motivo del rechazo.")
            return redirect("siniestro_detalle", pk=pk)

        siniestro.observaciones = f"RECHAZADO: {motivo}\n\n{siniestro.observaciones}"
        siniestro.estado = "rechazado"
        siniestro.fecha_cierre = timezone.now()
        siniestro.save(update_fields=["observaciones", "estado", "fecha_cierre"])

        messages.success(request, "Siniestro rechazado.")
        return redirect("siniestro_detalle", pk=pk)

    return render(
        request,
        "app/siniestros/rechazar.html",
        {
            "siniestro": siniestro,
        },
    )


def _notificar_custodio_inicio_proceso(siniestro):
    """
    Envía notificación al custodio informando que el proceso de siniestro ha iniciado.
    Se llama automáticamente cuando se notifica al broker.
    """
    from django.template.loader import render_to_string

    # Obtener custodio
    custodio = siniestro.responsable_custodio
    if not custodio:
        bien = siniestro.bien_asegurado
        if bien:
            custodio = bien.responsable_custodio

    if not custodio or not custodio.email:
        return  # No hay custodio o no tiene email

    try:
        # Datos del firmante
        firmante_nombre = ConfiguracionSistema.get_config("FIRMANTE_CARTA_NOMBRE", "Gestión de Seguros")
        firmante_cargo = ConfiguracionSistema.get_config("FIRMANTE_CARTA_CARGO", "Seguros / Inventarios")
        firmante_departamento = ConfiguracionSistema.get_config("FIRMANTE_DEPARTAMENTO", "")
        firmante_telefono = ConfiguracionSistema.get_config("FIRMANTE_TELEFONO", "")
        firmante_email = ConfiguracionSistema.get_config("FIRMANTE_EMAIL", "")

        # Datos del bien
        bien = siniestro.bien_asegurado
        bien_nombre = bien.nombre if bien else siniestro.get_nombre_bien()
        codigo_activo = bien.codigo_bien if bien else "N/A"

        # Nombre del custodio
        custodio_nombre = getattr(custodio, "nombre", None) or str(custodio)

        asunto = f"Proceso de Siniestro Iniciado - {siniestro.numero_siniestro}"

        # Contexto para la plantilla
        context = {
            "custodio_nombre": custodio_nombre,
            "numero_siniestro": siniestro.numero_siniestro,
            "bien_nombre": bien_nombre,
            "codigo_activo": codigo_activo,
            "fecha_reporte": siniestro.fecha_reporte.strftime("%d/%m/%Y") if siniestro.fecha_reporte else "N/A",
            "firmante_nombre": firmante_nombre,
            "firmante_cargo": firmante_cargo,
            "firmante_departamento": firmante_departamento,
            "firmante_telefono": firmante_telefono,
            "firmante_email": firmante_email,
            "timestamp": timezone.now().isoformat(),
        }

        # Renderizar plantilla HTML
        contenido_html = render_to_string("emails/inicio_proceso_custodio.html", context)

        # Versión texto plano
        contenido_texto = f"""Estimado/a {custodio_nombre},

Le informamos que se ha iniciado el proceso de gestión del siniestro relacionado con el bien bajo su custodia.

No. Siniestro: {siniestro.numero_siniestro}
Bien Afectado: {bien_nombre}
Código de Activo: {codigo_activo}

El corredor de seguros ha sido notificado y gestionará el reclamo ante la aseguradora.
Le mantendremos informado sobre el avance del proceso.

Atentamente,
{firmante_nombre}
{firmante_cargo}
"""

        email = _crear_email_con_logo(
            asunto=asunto,
            contenido_texto=contenido_texto,
            contenido_html=contenido_html,
            destinatario=custodio.email,
        )
        email.send()

        # Actualizar fecha de notificación al responsable
        siniestro.fecha_notificacion_responsable = timezone.now().date()
        siniestro.save(update_fields=["fecha_notificacion_responsable"])

    except Exception as e:
        import logging

        logging.error(f"Error enviando notificación al custodio: {e}")


def _enviar_notificaciones_cierre(siniestro):
    """
    Envía notificaciones de cierre al responsable y gerencia administrativa.
    """
    from django.template.loader import render_to_string

    asunto = f"Siniestro Cerrado - {siniestro.numero_siniestro}"

    # Obtener datos
    bien = siniestro.bien_asegurado
    bien_nombre = bien.nombre if bien else siniestro.get_nombre_bien()
    codigo_activo = bien.codigo_bien if bien else "N/A"
    poliza = siniestro.get_poliza()
    numero_poliza = poliza.numero_poliza if poliza else "N/A"

    # Datos del firmante
    firmante_nombre = ConfiguracionSistema.get_config("FIRMANTE_CARTA_NOMBRE", "Gestión de Seguros")
    firmante_cargo = ConfiguracionSistema.get_config("FIRMANTE_CARTA_CARGO", "Seguros / Inventarios")
    firmante_departamento = ConfiguracionSistema.get_config("FIRMANTE_DEPARTAMENTO", "")
    firmante_telefono = ConfiguracionSistema.get_config("FIRMANTE_TELEFONO", "")
    firmante_email = ConfiguracionSistema.get_config("FIRMANTE_EMAIL", "")

    # Lista de destinatarios con nombres
    destinatarios_info = []

    # Responsable/Custodio
    if siniestro.responsable_custodio and siniestro.responsable_custodio.email:
        nombre = getattr(siniestro.responsable_custodio, "nombre", None) or str(siniestro.responsable_custodio)
        destinatarios_info.append({"email": siniestro.responsable_custodio.email, "nombre": nombre})

    # Gerencia Administrativa
    email_gerencia = ConfiguracionSistema.get_config("EMAIL_GERENCIA_ADMINISTRATIVA", "")
    if email_gerencia:
        destinatarios_info.append({"email": email_gerencia, "nombre": "Gerencia Administrativa"})

    # Gerente de Siniestros
    email_gerente = ConfiguracionSistema.get_config("EMAIL_GERENTE_SINIESTROS", "")
    if email_gerente:
        destinatarios_info.append({"email": email_gerente, "nombre": "Gerente de Siniestros"})

    if destinatarios_info:
        try:
            for dest_info in destinatarios_info:
                # Contexto para la plantilla
                context = {
                    "numero_siniestro": siniestro.numero_siniestro,
                    "destinatario_nombre": dest_info["nombre"],
                    "bien_nombre": bien_nombre,
                    "codigo_activo": codigo_activo,
                    "numero_poliza": numero_poliza,
                    "fecha_cierre": siniestro.fecha_cierre.strftime("%d/%m/%Y") if siniestro.fecha_cierre else "N/A",
                    "monto_liquidado": f"{siniestro.monto_liquidado or 0:,.2f}",
                    "numero_comprobante": siniestro.numero_comprobante or "",
                    "firmante_nombre": firmante_nombre,
                    "firmante_cargo": firmante_cargo,
                    "firmante_departamento": firmante_departamento,
                    "firmante_telefono": firmante_telefono,
                    "firmante_email": firmante_email,
                    "timestamp": timezone.now().isoformat(),
                }

                # Renderizar plantilla HTML
                contenido_html = render_to_string("emails/cierre_siniestro.html", context)

                # Versión texto plano
                contenido_texto = f"""Estimado/a {dest_info['nombre']},

Se ha completado el proceso de indemnización del siniestro {siniestro.numero_siniestro}.

Bien: {bien_nombre}
Código de Activo: {codigo_activo}
Monto Liquidado: ${siniestro.monto_liquidado or 0:,.2f}
Comprobante: {siniestro.numero_comprobante or 'N/A'}
Fecha de Cierre: {context['fecha_cierre']}

Atentamente,
{firmante_nombre}
{firmante_cargo}
"""

                email = _crear_email_con_logo(
                    asunto=asunto,
                    contenido_texto=contenido_texto,
                    contenido_html=contenido_html,
                    destinatario=dest_info["email"],
                )
                email.send()
        except Exception:
            # logger.error(f"Error enviando notificaciones de cierre: {e}")
            pass


# ==============================================================================
# VISTAS DE CONFIGURACIÓN DEL SISTEMA
# ==============================================================================


@login_required
def configuracion_lista(request):
    """
    Lista todas las configuraciones del sistema agrupadas por categoría.
    Solo accesible para usuarios staff.
    """
    if not request.user.is_staff:
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    # Inicializar valores default si no existen
    ConfiguracionSistema.inicializar_valores_default()

    # Agrupar configuraciones por categoría
    categorias = {}
    categoria_labels = dict(ConfiguracionSistema._meta.get_field("categoria").choices)

    for config in ConfiguracionSistema.objects.all():
        cat = config.categoria
        if cat not in categorias:
            categorias[cat] = {"nombre": categoria_labels.get(cat, cat), "items": []}
        categorias[cat]["items"].append(config)

    # Ordenar categorías
    categorias_ordenadas = dict(sorted(categorias.items()))

    context = {
        "categorias": categorias_ordenadas,
        "titulo": "Configuración del Sistema",
    }

    return render(request, "app/configuracion/lista.html", context)


@login_required
def configuracion_editar(request, pk):
    """
    Edita una configuración individual del sistema.
    Solo accesible para usuarios staff.
    """
    if not request.user.is_staff:
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    config = get_object_or_404(ConfiguracionSistema, pk=pk)

    if request.method == "POST":
        form = ConfiguracionSistemaForm(request.POST, instance=config)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Configuración "{config.clave}" actualizada correctamente.')
                return redirect("configuracion_lista")
            except Exception as e:
                messages.error(request, f"Error al guardar: {str(e)}")
    else:
        form = ConfiguracionSistemaForm(instance=config)

    context = {
        "form": form,
        "config": config,
        "titulo": f"Editar: {config.clave}",
    }

    return render(request, "app/configuracion/editar.html", context)


@login_required
def configuracion_categoria(request, categoria):
    """
    Edita todas las configuraciones de una categoría en un solo formulario.
    Solo accesible para usuarios staff.
    """
    if not request.user.is_staff:
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    categoria_labels = dict(ConfiguracionSistema._meta.get_field("categoria").choices)
    categoria_nombre = categoria_labels.get(categoria, categoria)

    if request.method == "POST":
        form = ConfiguracionBulkForm(request.POST, categoria=categoria)
        if form.is_valid():
            try:
                saved = form.save()
                if saved:
                    messages.success(request, f'Se actualizaron {len(saved)} configuraciones: {", ".join(saved)}')
                else:
                    messages.info(request, "No hubo cambios en las configuraciones.")
                return redirect("configuracion_lista")
            except Exception as e:
                messages.error(request, f"Error al guardar: {str(e)}")
    else:
        form = ConfiguracionBulkForm(categoria=categoria)

    context = {
        "form": form,
        "categoria": categoria,
        "categoria_nombre": categoria_nombre,
        "titulo": f"Configuración: {categoria_nombre}",
    }

    return render(request, "app/configuracion/categoria.html", context)


@login_required
def configuracion_restablecer(request):
    """
    Restablece todas las configuraciones a sus valores predeterminados.
    Solo accesible para superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden restablecer la configuración.")
        return redirect("configuracion_lista")

    if request.method == "POST":
        # Eliminar todas las configuraciones actuales
        ConfiguracionSistema.objects.all().delete()
        # Reinicializar con valores default
        ConfiguracionSistema.inicializar_valores_default()
        messages.success(request, "Todas las configuraciones han sido restablecidas a sus valores predeterminados.")
        return redirect("configuracion_lista")

    return render(
        request,
        "app/configuracion/restablecer.html",
        {
            "titulo": "Restablecer Configuración",
        },
    )


# ==============================================================================
# VISTAS DE RESPALDO Y RECUPERACIÓN
# ==============================================================================


@login_required
def backups_lista(request):
    """
    Lista todos los backups del sistema con estadísticas.
    Solo accesible para usuarios staff.
    """
    if not request.user.is_staff:
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    # Obtener backups
    backups = BackupRegistro.objects.all()[:50]

    # Filtros
    tipo_filtro = request.GET.get("tipo", "")
    estado_filtro = request.GET.get("estado", "")

    if tipo_filtro:
        backups = backups.filter(tipo=tipo_filtro)
    if estado_filtro:
        backups = backups.filter(estado=estado_filtro)

    # Estadísticas
    stats = BackupRegistro.obtener_estadisticas()
    config = ConfiguracionBackup.get_config()

    context = {
        "backups": backups,
        "stats": stats,
        "config": config,
        "tipo_filtro": tipo_filtro,
        "estado_filtro": estado_filtro,
        "tipos": BackupRegistro.TIPO_CHOICES,
        "estados": BackupRegistro.ESTADO_CHOICES,
        "titulo": "Respaldos del Sistema",
    }

    return render(request, "app/backups/lista.html", context)


@login_required
def backup_crear(request):
    """
    Crea un nuevo backup manual del sistema.
    Solo accesible para usuarios superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden crear backups.")
        return redirect("backups_lista")

    if request.method == "POST":
        import time

        from django.core.management import call_command

        incluir_media = request.POST.get("incluir_media") == "on"
        comprimir = request.POST.get("comprimir", "on") == "on"
        notas = request.POST.get("notas", "")

        # Crear registro
        backup = BackupRegistro.objects.create(
            nombre=f'backup_manual_{timezone.now().strftime("%Y%m%d_%H%M%S")}',
            ruta="",
            tipo="completo" if incluir_media else "base_datos",
            estado="en_progreso",
            frecuencia="manual",
            comprimido=comprimir,
            creado_por=request.user,
            notas=notas,
        )

        try:
            inicio = time.time()

            # Ejecutar backup usando StringIO para capturar salida
            from io import StringIO

            out = StringIO()

            # Construir argumentos
            args = ["backup_database"]
            kwargs = {
                "stdout": out,
                "verbosity": 0,
            }
            if comprimir:
                kwargs["compress"] = True
            if incluir_media:
                kwargs["include_media"] = True

            call_command(*args, **kwargs)
            resultado = out.getvalue().strip()

            # Actualizar registro
            from pathlib import Path

            duracion = int(time.time() - inicio)
            backup_path = Path(resultado) if resultado else None

            backup.estado = "completado"
            backup.duracion_segundos = duracion
            if backup_path and backup_path.exists():
                backup.ruta = str(backup_path)
                backup.tamaño = backup_path.stat().st_size
                backup.nombre = backup_path.name
            backup.save()

            messages.success(request, f"Backup creado exitosamente: {backup.nombre} ({backup.tamaño_legible})")

        except Exception as e:
            backup.estado = "fallido"
            backup.error_mensaje = str(e)
            backup.save()
            messages.error(request, f"Error al crear backup: {str(e)}")

        return redirect("backups_lista")

    context = {
        "titulo": "Crear Nuevo Backup",
    }

    return render(request, "app/backups/crear.html", context)


@login_required
def backup_descargar(request, pk):
    """
    Descarga un archivo de backup.
    Solo accesible para usuarios superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden descargar backups.")
        return redirect("backups_lista")

    from pathlib import Path

    from django.http import FileResponse

    backup = get_object_or_404(BackupRegistro, pk=pk)

    archivo = Path(backup.ruta)
    if not archivo.exists():
        messages.error(request, "El archivo de backup no existe.")
        return redirect("backups_lista")

    response = FileResponse(open(archivo, "rb"), as_attachment=True, filename=backup.nombre)

    return response


@login_required
def backup_eliminar(request, pk):
    """
    Elimina un backup del sistema.
    Solo accesible para usuarios superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden eliminar backups.")
        return redirect("backups_lista")

    backup = get_object_or_404(BackupRegistro, pk=pk)

    if request.method == "POST":
        from pathlib import Path

        # Eliminar archivo físico
        archivo = Path(backup.ruta)
        if archivo.exists():
            archivo.unlink()

        # Marcar como eliminado
        backup.estado = "eliminado"
        backup.save()

        messages.success(request, f'Backup "{backup.nombre}" eliminado correctamente.')
        return redirect("backups_lista")

    return render(
        request,
        "app/backups/eliminar.html",
        {
            "backup": backup,
            "titulo": f"Eliminar Backup: {backup.nombre}",
        },
    )


@login_required
def backup_restaurar(request, pk):
    """
    Restaura el sistema desde un backup.
    Solo accesible para usuarios superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden restaurar backups.")
        return redirect("backups_lista")

    backup = get_object_or_404(BackupRegistro, pk=pk)

    if not backup.archivo_existe:
        messages.error(request, "El archivo de backup no existe.")
        return redirect("backups_lista")

    if request.method == "POST":
        from django.core.management import call_command

        confirmacion = request.POST.get("confirmacion", "")
        if confirmacion != "RESTAURAR":
            messages.error(request, "Debe escribir RESTAURAR para confirmar.")
            return redirect("backup_restaurar", pk=pk)

        try:
            # Ejecutar restauración
            call_command(
                "restore_database", backup.ruta, compressed=backup.comprimido, no_confirm=True, backup_first=True
            )

            # Registrar restauración
            BackupRegistro.objects.create(
                nombre=f"restauracion_desde_{backup.nombre}",
                ruta=backup.ruta,
                tipo="restauracion",
                estado="completado",
                creado_por=request.user,
                notas=f"Restauración desde backup {backup.nombre}",
            )

            messages.success(request, "Sistema restaurado exitosamente.")
            return redirect("backups_lista")

        except Exception as e:
            messages.error(request, f"Error al restaurar: {str(e)}")
            return redirect("backup_restaurar", pk=pk)

    context = {
        "backup": backup,
        "titulo": f"Restaurar desde: {backup.nombre}",
    }

    return render(request, "app/backups/restaurar.html", context)


@login_required
def backup_configuracion(request):
    """
    Configura los backups automáticos del sistema.
    Solo accesible para usuarios superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden configurar backups.")
        return redirect("backups_lista")

    config = ConfiguracionBackup.get_config()

    if request.method == "POST":
        config.activo = request.POST.get("activo") == "on"
        config.frecuencia = request.POST.get("frecuencia", "diario")
        config.hora_ejecucion = request.POST.get("hora_ejecucion", "02:00")
        config.dias_retener = int(request.POST.get("dias_retener", 30))
        config.incluir_media = request.POST.get("incluir_media") == "on"
        config.comprimir = request.POST.get("comprimir", "on") == "on"
        config.notificar_email = request.POST.get("notificar_email", "")
        config.save()

        messages.success(request, "Configuración de backups actualizada.")
        return redirect("backups_lista")

    context = {
        "config": config,
        "frecuencias": BackupRegistro.FRECUENCIA_CHOICES,
        "titulo": "Configuración de Backups Automáticos",
    }

    return render(request, "app/backups/configuracion.html", context)
